""" Test schema IO

:Author: Jonathan Karr <karr@mssm.edu>
:Author: Arthur Goldberg <Arthur.Goldberg@mssm.edu>
:Date: 2016-11-23
:Copyright: 2016, Karr Lab
:License: MIT
"""

from os.path import splitext
from obj_tables import core, utils, chem, ontology, units
from obj_tables.io import WorkbookReader, WorkbookWriter, convert, create_template, IoWarning
from pathlib import Path
from wc_utils.workbook.io import (Workbook, Worksheet, Row, WorkbookStyle, WorksheetStyle,
                                  read as read_workbook, write as write_workbook, get_reader, get_writer)
import datetime
import git
import enum
import json
import math
import mock
import obj_tables
import obj_tables.io
import obj_tables.expression
import openpyxl
import os
import pint
import pronto
import pytest
import re
import shutil
import sys
import tempfile
import unittest
import warnings
import wc_utils.util.chem
from wc_utils.util.git import GitHubRepoForTests


class MainRoot(core.Model):
    id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')
    name = core.StringAttribute()

    class Meta(core.Model.Meta):
        attribute_order = ('id', 'name', )
        table_format = core.TableFormat.column


class Node(core.Model):
    id = core.SlugAttribute(primary=True)
    root = core.ManyToOneAttribute(MainRoot, related_name='nodes')
    val1 = core.FloatAttribute()
    val2 = core.FloatAttribute()

    class Meta(core.Model.Meta):
        attribute_order = ('id', 'root', 'val1', 'val2', )
        indexed_attrs_tuples = (('id',), ('id', 'val1'))


class OneToManyRowAttribute(core.OneToManyAttribute):
    pass


class OneToManyInlineAttribute(core.OneToManyAttribute):

    def serialize(self, value, encoded=None):
        return ', '.join([obj.id for obj in value])

    def deserialize(self, value, objects, decoded=None):
        if value:
            objs = []
            for id in value.split(', '):
                obj = OneToManyInline(id=id)
                if OneToManyInline not in objects:
                    objects[OneToManyInline] = {}
                objects[OneToManyInline][id] = obj
                objs.append(obj)

            return (objs, None)
        else:
            return (set(), None)


class Leaf(core.Model):
    id = core.StringAttribute(primary=True)
    nodes = core.ManyToManyAttribute(Node, related_name='leaves')
    val1 = core.FloatAttribute()
    val2 = core.FloatAttribute()
    onetomany_rows = OneToManyRowAttribute('OneToManyRow', related_name='leaf', min_related_rev=1)
    onetomany_inlines = OneToManyInlineAttribute('OneToManyInline', related_name='leaf', min_related_rev=1)

    class Meta(core.Model.Meta):
        attribute_order = (
            'id', 'nodes', 'val1', 'val2',
            'onetomany_rows', 'onetomany_inlines'
        )


class OneToManyRow(core.Model):
    id = core.SlugAttribute(primary=True)

    class Meta(core.Model.Meta):
        attribute_order = ('id',)


class OneToManyInline(core.Model):
    id = core.SlugAttribute(primary=False)

    class Meta(core.Model.Meta):
        attribute_order = ('id',)
        table_format = core.TableFormat.cell


class TestIo(unittest.TestCase):

    def setUp(self):
        self.root = root = MainRoot(id='root', name=u'\u20ac')
        nodes = [
            Node(root=root, id='node_0', val1=1, val2=2),
            Node(root=root, id='node_1', val1=3, val2=4),
            Node(root=root, id='node_2', val1=5, val2=6),
        ]
        self.leaves = leaves = [
            Leaf(nodes=[nodes[0]], id='leaf_0_0', val1=7, val2=8),
            Leaf(nodes=[nodes[0]], id='leaf_0_1', val1=9, val2=10),
            Leaf(nodes=[nodes[1]], id='leaf_1_0', val1=11, val2=12),
            Leaf(nodes=[nodes[1]], id='leaf_1_1', val1=13, val2=14),
            Leaf(nodes=[nodes[2]], id='leaf_2_0', val1=15, val2=16),
            Leaf(nodes=[nodes[2]], id='leaf_2_1', val1=17, val2=18),
        ]
        leaves[0].onetomany_rows = [OneToManyRow(id='row_0_0'), OneToManyRow(id='row_0_1')]
        leaves[1].onetomany_rows = [OneToManyRow(id='row_1_0'), OneToManyRow(id='row_1_1')]
        leaves[2].onetomany_rows = [OneToManyRow(id='row_2_0'), OneToManyRow(id='row_2_1')]
        leaves[3].onetomany_rows = [OneToManyRow(id='row_3_0'), OneToManyRow(id='row_3_1')]
        leaves[4].onetomany_rows = [OneToManyRow(id='row_4_0'), OneToManyRow(id='row_4_1')]
        leaves[5].onetomany_rows = [OneToManyRow(id='row_5_0'), OneToManyRow(id='row_5_1')]

        leaves[0].onetomany_inlines = [OneToManyInline(id='inline_0_0'), OneToManyInline(id='inline_0_1')]
        leaves[1].onetomany_inlines = [OneToManyInline(id='inline_1_0'), OneToManyInline(id='inline_1_1')]
        leaves[2].onetomany_inlines = [OneToManyInline(id='inline_2_0'), OneToManyInline(id='inline_2_1')]
        leaves[3].onetomany_inlines = [OneToManyInline(id='inline_3_0'), OneToManyInline(id='inline_3_1')]
        leaves[4].onetomany_inlines = [OneToManyInline(id='inline_4_0'), OneToManyInline(id='inline_4_1')]
        leaves[5].onetomany_inlines = [OneToManyInline(id='inline_5_0'), OneToManyInline(id='inline_5_1')]

        self.tmp_dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp_dirname)

    def test_dummy_model(self):
        # test integrity of relationships
        for leaf in self.leaves:
            for row in leaf.onetomany_rows:
                self.assertEqual(row.leaf, leaf)

    def test_write_read(self):
        # write/read to/from Excel
        root = self.root
        objects = list(set([root] + root.get_related()))
        objects = utils.group_objects_by_model(objects)

        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [root], models=[MainRoot, Node, Leaf, ])
        WorkbookWriter().run(filename, root, models=[MainRoot, Node, Leaf, ])
        objects2 = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf, OneToManyRow])

        # validate
        all_objects2 = []
        for model, model_objects in objects2.items():
            all_objects2.extend(model_objects)
        self.assertEqual(core.Validator().run(all_objects2), None)

        # test objects saved and loaded correctly
        for model in objects.keys():
            self.assertEqual(len(objects2[model]), len(objects[model]),
                             msg='Different numbers of "{}" objects'.format(model.__name__))
        self.assertEqual(len(objects2), len(objects))

        root2 = objects2[MainRoot].pop()

        filename2 = os.path.join(self.tmp_dirname, 'test2.xlsx')
        WorkbookWriter().run(filename2, [root2], models=[MainRoot, Node, Leaf, ])
        original = read_workbook(filename)
        copy = read_workbook(filename2)
        for sheet in copy.keys():
            copy[sheet][0][0] = original[sheet][0][0]  # because dates will be different
        self.assertEqual(copy, original)

        self.assertEqual(set([x.id for x in root2.nodes]), set([x.id for x in root.nodes]))
        self.assertTrue(root2.is_equal(root))

        # unicode
        self.assertEqual(root2.name, u'\u20ac')

        #
        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        WorkbookWriter().run(filename, root, models=[MainRoot, Node, Leaf, ])

        WorkbookWriter().run(filename, None, models=[MainRoot, Node, Leaf, ])
        objects2 = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf, ], group_objects_by_model=False)
        self.assertEqual(objects2, None)
        objects2 = obj_tables.io.Reader().run(filename, models=[MainRoot, Node, Leaf, ], group_objects_by_model=False)
        self.assertEqual(objects2, None)

    def test_write_read_2(self):
        # test obj_tables.io.Writer, obj_tables.io.Reader
        root = self.root
        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        obj_tables.io.Writer().run(filename, root, models=[MainRoot, Node, Leaf, OneToManyRow])
        objects2 = obj_tables.io.Reader().run(filename, models=[MainRoot, Node, Leaf, OneToManyRow],
                                              group_objects_by_model=True)
        self.assertTrue(root.is_equal(objects2[MainRoot][0]))

        # test no validation
        wb = read_workbook(filename)
        wb['!Nodes'][4][2] = 'node_0'
        filename2 = os.path.join(self.tmp_dirname, 'test2.xlsx')
        write_workbook(filename2, wb, style={
            MainRoot.Meta.verbose_name: WorksheetStyle(extra_rows=0, extra_columns=0),
            Node.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            Leaf.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            OneToManyRow.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaises(ValueError):
            WorkbookReader().run(filename2, models=[MainRoot, Node, Leaf, OneToManyRow],
                                 group_objects_by_model=True, validate=True)
        WorkbookReader().run(filename2, models=[MainRoot, Node, Leaf, OneToManyRow],
                             group_objects_by_model=True, validate=False)

    def test_manager(self):

        class Example0(core.Model):
            id = core.SlugAttribute(primary=True)
            int_attr = core.IntegerAttribute()

            class Meta(core.Model.Meta):
                indexed_attrs_tuples = (('id',),)
                attribute_order = ('id', 'int_attr')

        class Example1(core.Model):
            str_attr = core.StringAttribute()
            int_attr = core.IntegerAttribute()
            int_attr2 = core.IntegerAttribute()
            test0 = core.OneToOneAttribute(Example0, related_name='test1')

            class Meta(core.Model.Meta):
                indexed_attrs_tuples = (('str_attr',), ('int_attr', 'int_attr2'),)
                attribute_order = ('str_attr', 'int_attr', 'int_attr2', 'test0')

        filename = os.path.join(os.path.dirname(__file__), 'fixtures', 'test_manager.xlsx')
        WorkbookReader().run(filename, models=[Example0, Example1])

        self.assertEqual(len(Example0.objects.get(id='A')), 1)
        self.assertEqual(len(Example1.objects.get(str_attr='C')), 2)
        self.assertEqual(len(Example1.objects.get(int_attr=11, int_attr2=21)), 1)
        self.assertEqual(Example0.objects.get_one(id='A'),
                         Example1.objects.get_one(int_attr=11, int_attr2=21).test0)
        with self.assertRaises(ValueError) as context:
            Example0.objects.get(int_attr=1)
        self.assertIn("not an indexed attribute tuple", str(context.exception))

    def test_read_inexact_worksheet_name_match(self):
        filename = os.path.join(self.tmp_dirname, 'test-*.csv')

        # write to file
        WorkbookWriter().run(filename, [self.root], models=[MainRoot, Node, Leaf, ])

        """ test reading worksheet by the model's name """
        # rename worksheet
        self.assertTrue(os.path.isfile(os.path.join(self.tmp_dirname, 'test-Main root.csv')))
        os.rename(os.path.join(self.tmp_dirname, 'test-Main root.csv'), os.path.join(self.tmp_dirname, 'test-MainRoot.csv'))

        objects = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf, OneToManyRow])
        root = objects[MainRoot].pop()

        self.assertTrue(root.is_equal(self.root))

        """ test reading worksheet by the model's verbose name """
        # rename worksheet
        self.assertTrue(os.path.isfile(os.path.join(self.tmp_dirname, 'test-Leaves.csv')))
        os.rename(os.path.join(self.tmp_dirname, 'test-Leaves.csv'), os.path.join(self.tmp_dirname, 'test-Leaf.csv'))

        objects = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf, OneToManyRow])
        root = objects[MainRoot].pop()

        self.assertTrue(root.is_equal(self.root))

        """ test reading worksheet by the model's plural verbose name """
        # rename worksheet
        self.assertTrue(os.path.isfile(os.path.join(self.tmp_dirname, 'test-MainRoot.csv')))
        os.rename(os.path.join(self.tmp_dirname, 'test-MainRoot.csv'), os.path.join(self.tmp_dirname, 'test-Main roots.csv'))

        objects = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf, OneToManyRow])
        root = objects[MainRoot].pop()

        self.assertTrue(root.is_equal(self.root))

        """ test reading worksheet by the model's plural verbose name, case-insensitive """
        # rename worksheet
        self.assertTrue(os.path.isfile(os.path.join(self.tmp_dirname, 'test-Main roots.csv')))
        os.rename(os.path.join(self.tmp_dirname, 'test-Main roots.csv'), os.path.join(self.tmp_dirname, 'test-main roots.csv'))

        objects = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf, OneToManyRow])
        root = objects[MainRoot].pop()

        self.assertTrue(root.is_equal(self.root))

    def test_read_inexact_attribute_name_match(self):
        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        filename2 = os.path.join(self.tmp_dirname, 'test2.xlsx')

        # write to file
        WorkbookWriter().run(filename, [self.root], models=[MainRoot, Node, Leaf, ])

        """ test reading attributes by verbose name """
        objects = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf, OneToManyRow])
        root = objects[MainRoot].pop()

        self.assertTrue(root.is_equal(self.root))

        """ test reading attributes by name """
        # setup reader, writer
        _, ext = splitext(filename)
        reader_cls = get_reader(ext)
        writer_cls = get_writer(ext)
        reader = reader_cls(filename)
        writer = writer_cls(filename2)

        # read workbook
        workbook = reader.run()

        # edit heading
        headings = workbook['!Main root'][1]
        self.assertEqual(headings[0], '!Identifier')
        headings[0] = '!id'

        # write workbook
        writer.run(workbook, style={
            MainRoot.Meta.verbose_name: WorksheetStyle(extra_rows=0, extra_columns=0),
            Node.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            Leaf.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            OneToManyRow.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })

        # check that attributes can be read by name
        objects = WorkbookReader().run(filename2, models=[MainRoot, Node, Leaf, OneToManyRow])
        root = objects[MainRoot].pop()

        self.assertTrue(root.is_equal(self.root))

        """ test case insensitivity """
        # edit heading
        workbook['!Main root'][1][0] = '!ID'

        # write workbook
        writer.run(workbook, style={
            MainRoot.Meta.verbose_name: WorksheetStyle(extra_rows=0, extra_columns=0),
            Node.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            Leaf.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            OneToManyRow.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })

        # check that attributes can be read by name
        objects = WorkbookReader().run(filename2, models=[MainRoot, Node, Leaf, OneToManyRow])
        root = objects[MainRoot].pop()

        self.assertTrue(root.is_equal(self.root))

    def test_validation(self):
        t = MainRoot(name='f')
        self.assertIn('value for primary attribute cannot be empty',
                      t.validate().attributes[0].messages[0])

    def check_reader_errors(self, fixture_file, expected_messages, models, use_re=False,
                            do_not_catch=False):
        ''' Run WorkbookReader expecting an error; check that the exception message matches expected messages

        Args:
            fixture_file (:obj:`str`): name of the file to be read
            expected_messages (:obj:`list` of :obj:`str`): list of expected strings or patterns in the
                exception
            models (:obj:`list` of :obj:`Model`): :obj:`Model`\ s for the schema of the data being read
            use_re (:obj:`boolean`, optional): if set, `expected_messages` contains RE patterns
            do_not_catch (:obj:`boolean`, optional): if set, run WorkbookReader() outside try ... catch;
                produces full exception message for debugging

        Raises:
            :obj:`Exception`: if do_not_catch
        '''
        filename = os.path.join(os.path.dirname(__file__), 'fixtures', fixture_file)
        if do_not_catch:
            WorkbookReader().run(filename, models=models)
        with self.assertRaises(Exception) as context:
            WorkbookReader().run(filename, models=models)
        for msg in expected_messages:
            if not use_re:
                msg = re.escape(msg)
            self.assertRegex(str(context.exception), msg)

    def test_location_of_attrs(self):
        class Normal(core.Model):
            id = core.SlugAttribute()
            val = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'val')

        class Transposed(core.Model):
            tid = core.SlugAttribute()
            s = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('tid', 's', )
                table_format = core.TableFormat.column

        file = 'test-locations.xlsx'
        filename = os.path.join(os.path.dirname(__file__), 'fixtures', file)
        models = WorkbookReader().run(filename, models=[Normal, Transposed])
        ext = 'xlsx'
        normals = models[Normal]
        for obj in normals:
            if obj.val == 'x':
                (file_type, basename, worksheet, row, column) = obj.get_source('val')
                self.assertEqual(file_type, ext)
                self.assertEqual(basename, file)
                self.assertEqual(worksheet, '!' + obj.Meta.verbose_name_plural)
                self.assertEqual(row, 3)
                self.assertEqual(column, 'B')
                self.assertEqual(utils.source_report(obj, 'val'),
                                 ':'.join([file, '!' + obj.Meta.verbose_name_plural, "{}{}".format(column, row)]))

        transposeds = models[Transposed]
        for obj in transposeds:
            if obj.s == 'z':
                (file_type, basename, worksheet, row, column) = obj.get_source('s')
                self.assertEqual(file_type, ext)
                self.assertEqual(basename, file)
                self.assertEqual(worksheet, '!' + obj.Meta.verbose_name)
                self.assertEqual(row, 2)
                self.assertEqual(column, 'C')
                self.assertEqual(utils.source_report(obj, 's'),
                                 ':'.join([file, '!' + obj.Meta.verbose_name, "{}{}".format(column, row)]))

        file = 'test-locations-*.csv'
        filename = os.path.join(os.path.dirname(__file__), 'fixtures', file)
        models = WorkbookReader().run(filename, models=[Normal, Transposed])
        ext = 'csv'
        normals = models[Normal]
        for obj in normals:
            if obj.val == 'x':
                (file_type, basename, worksheet, row, column) = obj.get_source('val')
                self.assertEqual(file_type, ext)
                self.assertEqual(basename, file)
                self.assertEqual(row, 3)
                self.assertEqual(worksheet, obj.Meta.verbose_name_plural)
                self.assertEqual(column, 2)
                self.assertEqual(utils.source_report(obj, 'val'),
                                 ':'.join([file, obj.Meta.verbose_name_plural, "{},{}".format(row, column)]))

        transposeds = models[Transposed]
        for obj in transposeds:
            if obj.s == 'z':
                (file_type, basename, worksheet, row, column) = obj.get_source('s')
                self.assertEqual(file_type, ext)
                self.assertEqual(basename, file)
                self.assertEqual(worksheet, obj.Meta.verbose_name)
                self.assertEqual(row, 2)
                self.assertEqual(column, 3)
                self.assertEqual(utils.source_report(obj, 's'),
                                 ':'.join([file, obj.Meta.verbose_name, "{},{}".format(row, column)]))

    def test_read_bad_headers(self):
        msgs = [
            "The model cannot be loaded because 'bad-headers.xlsx' contains error(s)",
            "Header 'y' in row 1, col F does not match any attribute",
        ]
        self.check_reader_errors('bad-headers.xlsx', msgs, [MainRoot, Node, Leaf, OneToManyRow])

        msgs = [
            "The model cannot be loaded because 'bad-headers-*.csv' contains error(s)",
            "Header 'x' in row 5, col 1 does not match any attribute",
        ]
        self.check_reader_errors('bad-headers-*.csv', msgs, [MainRoot, Node, Leaf, OneToManyRow])

        '''
        msgs = [
            "Duplicate, case insensitive, header fields: 'MainRoot', 'root'",
            "Duplicate, case insensitive, header fields: 'good val', 'Good val', 'Good VAL'"]
        self.check_reader_errors('duplicate-headers.xlsx', msgs, [Node])
        '''

    def test_uncaught_data_error(self):
        class Test(core.Model):
            id = core.SlugAttribute(primary=True)
            float1 = core.FloatAttribute()
            bool1 = core.FloatAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'float1', 'bool1', )

        msgs = ["The model cannot be loaded because 'uncaught-error.xlsx' contains error(s)",
                "uncaught-error.xlsx:!Tests:B5",
                "uncaught-error.xlsx:!Tests:C6",
                "Value must be an instance of `float`",
                ]
        self.check_reader_errors('uncaught-error.xlsx', msgs, [MainRoot, Test])

    def test_read_invalid_data(self):
        class NormalRecord(core.Model):
            id_with_underscores = core.SlugAttribute()
            val = core.StringAttribute(min_length=2)

            class Meta(core.Model.Meta):
                attribute_order = ('id_with_underscores', 'val')

        class Transposed(core.Model):
            id = core.SlugAttribute()
            val = core.StringAttribute(min_length=2)

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'val', )
                table_format = core.TableFormat.column

        RE_msgs = [
            "Leaf\n"
            " +:\n"
            " +'id':''\n"
            " +invalid-data.xlsx:!Leaves:A6\n"
            " +StringAttribute value for primary attribute cannot be empty",
            "invalid-data.xlsx:'!Normal records':B3",
            "Transposed\n"
            " +t_2:\n"
            " +'val':'x'\n"
            " +invalid-data.xlsx:!Transposed:C2\n"
            " +Value must be at least 2 characters",
        ]
        self.check_reader_errors('invalid-data.xlsx', RE_msgs, [Leaf, NormalRecord, Transposed],
                                 use_re=True)

        RE_msgs = [
            r"The model cannot be loaded because 'invalid-data-\*.csv' contains error",
            r"Leaf *\n +:\n +'id':''\n +invalid-data-\*.csv:Leaves:6,1\n +StringAttribute value for "
            r"primary attribute cannot be empty",
            r"Transposed\n +t_2:\n +'val':'x'\n +invalid-data-\*.csv:Transposed:2,3\n +Value must be at "
            r"least 2 characters",
        ]
        self.check_reader_errors('invalid-data-*.csv', RE_msgs, [Leaf, NormalRecord, Transposed],
                                 use_re=True)

    def test_reference_errors(self):
        class NodeFriend(core.Model):
            id = core.SlugAttribute()
            node = core.OneToOneAttribute(Node, related_name='nodes')
            val = core.StringAttribute(min_length=2)

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'node', 'val')

        RE_msgs = [
            "reference-errors.xlsx:!Nodes:B3\n +Unable to find MainRoot with id='not root'",
            "reference-errors.xlsx:!Leaves:B6\n +Unable to find Node with id='no such node'",
            "reference-errors.xlsx:!Leaves:E7\n +Unable to find OneToManyRow with id='no such row'",
            "reference-errors.xlsx:'!Node friends':B2\n +Unable to find Node with id=no_node",
        ]
        self.check_reader_errors('reference-errors.xlsx', RE_msgs,
                                 [MainRoot, Node, NodeFriend, Leaf, OneToManyRow], use_re=True)

    def test_duplicate_primaries(self):
        RE_msgs = [
            "The model cannot be loaded because it fails to validate",
            "Node:\n +'id':\n +id values must be unique, but these values are repeated: node_2",
            "MainRoot:\n +'id':\n +id values must be unique, but these values are repeated: 'root 2'",
        ]
        self.check_reader_errors('duplicate-primaries.xlsx', RE_msgs, [MainRoot, Node, Leaf, OneToManyRow],
                                 use_re=True)

    def test_create_worksheet_style(self):
        self.assertIsInstance(WorkbookWriter.create_worksheet_style(MainRoot), WorksheetStyle)

    def test_convert(self):
        filename_xls1 = os.path.join(self.tmp_dirname, 'test1.xlsx')
        filename_xls2 = os.path.join(self.tmp_dirname, 'test2.xlsx')
        filename_csv = os.path.join(self.tmp_dirname, 'test-*.csv')

        models = [MainRoot, Node, Leaf, OneToManyRow]

        WorkbookWriter().run(filename_xls1, [self.root], models=models)

        convert(filename_xls1, filename_csv, models)
        convert(filename_csv, filename_xls2, models)

        objects2 = WorkbookReader().run(filename_csv, models=models)
        self.assertTrue(self.root.is_equal(objects2[MainRoot][0]))

        objects2 = WorkbookReader().run(filename_xls2, models=models)
        self.assertTrue(self.root.is_equal(objects2[MainRoot][0]))

    def test_create_template(self):
        filename = os.path.join(self.tmp_dirname, 'test3.xlsx')
        create_template(filename, [MainRoot, Node, Leaf])
        objects = WorkbookReader().run(filename, models=[MainRoot, Node, Leaf])
        self.assertEqual(objects, {
            MainRoot: [],
            Node: [],
            Leaf: [],
        })
        self.assertEqual(core.Validator().run([]), None)

    def run_options_helper(self, fixture_file):
        filename = os.path.join(os.path.dirname(__file__), 'fixtures', fixture_file)

        class SimpleModel(core.Model):
            val = core.StringAttribute(min_length=10)

        class ExtraSheet(core.Model):
            val = core.StringAttribute(min_length=10)

        with self.assertRaises(ValueError) as context:
            # raises extra sheet exception
            WorkbookReader().run(filename, models=[SimpleModel])
        self.assertEqual(str(context.exception),
                         "No matching models for worksheets with TableIds 'extra sheet' in {}".format(
            fixture_file))

        with self.assertRaises(ValueError) as context:
            # raises extra attribute exception
            WorkbookReader().run(filename, models=[SimpleModel, ExtraSheet])
        self.assertRegex(str(context.exception),
                         "The model cannot be loaded because 'test_run_options.*' contains error.*")
        if 'xlsx' in fixture_file:
            col = 'B'
        elif 'csv' in fixture_file:
            col = '2'
        self.assertRegex(str(context.exception),
                         ".*Header 'extra' in row 1, col {} does not match any attribute.*".format(col))

        with self.assertRaises(ValueError) as context:
            # raises validation exception on 'too short'
            WorkbookReader().run(filename, models=[SimpleModel, ExtraSheet],
                                 ignore_extra_attributes=True)
        self.assertRegex(str(context.exception),
                         "The model cannot be loaded because 'test_run_options.*' contains error.*")
        if 'xlsx' in fixture_file:
            location = 'A3'
        elif 'csv' in fixture_file:
            location = '3,1'
        if fixture_file.endswith('.xlsx'):
            prefix = '!'
        else:
            prefix = ''
        self.assertRegex(str(context.exception),
                         ".*'val':'too short'\n.*test_run_options.*:'{}Simple models':{}\n.*"
                         "Value must be at least 10 characters".format(prefix, location))

        class SimpleModel(core.Model):
            val = core.StringAttribute()
        model = WorkbookReader().run(filename, models=[SimpleModel, ExtraSheet],
                                     ignore_extra_attributes=True)
        self.assertIn('too short', [r.val for r in model[SimpleModel]])

    def test_run_options(self):
        self.run_options_helper('test_run_options.xlsx')
        self.run_options_helper('test_run_options-*.csv')

    def test_read_empty_numeric_cell(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet('!Test models')
        cell = ws.cell(row=1, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = "!!ObjTables TableType='Data' ModelId='TestModel'"

        cell = ws.cell(row=2, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = '!Id'

        cell = ws.cell(row=2, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = '!Value'

        cell = ws.cell(row=3, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'Model-1'

        cell = ws.cell(row=3, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
        cell.value = 2.

        cell = ws.cell(row=4, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'Model-2'

        cell = ws.cell(row=4, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
        cell.value = 3.

        cell = ws.cell(row=5, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'Model-3'

        cell = ws.cell(row=5, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
        cell.value = None

        cell = ws.cell(row=6, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'Model-4'

        cell = ws.cell(row=6, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = None

        cell = ws.cell(row=7, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'Model-5'

        cell = ws.cell(row=7, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = ''

        cell = ws.cell(row=8, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'Model-6'

        cell = ws.cell(row=8, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_NULL
        cell.value = None

        cell = ws.cell(row=9, column=1)
        cell.data_type = openpyxl.cell.cell.TYPE_STRING
        cell.value = 'Model-7'

        cell = ws.cell(row=9, column=2)
        cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
        cell.value = 5.

        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        wb.save(filename)

        class TestModel(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            value = core.FloatAttribute()

        models = WorkbookReader().run(filename, models=[TestModel])[TestModel]
        models.sort(key=lambda model: model.id)

        m = TestModel.objects.get_one(id='Model-2')
        self.assertEqual(models[0].value, 2.)
        self.assertEqual(models[1].value, 3.)
        self.assertTrue(math.isnan(models[2].value))
        self.assertTrue(math.isnan(models[3].value))
        self.assertTrue(math.isnan(models[4].value))
        self.assertTrue(math.isnan(models[5].value))
        self.assertEqual(models[6].value, 5.)

    def test_not_existant_referenced_object(self):
        class ErrorAttribute(core.StringAttribute):
            def deserialize(self, value):
                raise Exception()

        class TestModel(core.Model):
            id = ErrorAttribute(primary=True, unique=True)

        workbook = Workbook()

        workbook['!Test models'] = worksheet = Worksheet()
        worksheet.append(Row(["!!ObjTables TableType='Data' ModelId='TestModel'"]))
        worksheet.append(Row(['!Id']))
        worksheet.append(Row(['A']))
        worksheet.append(Row(['B']))
        worksheet.append(Row(['C']))

        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        xslx_writer = get_writer('.xlsx')(filename)
        xslx_writer.run(workbook, style={
            TestModel.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })

        with self.assertRaisesRegex(ValueError, 'The model cannot be loaded'):
            WorkbookReader().run(filename, models=[TestModel])

    def test_table_format_multiple_cells(self):
        class Unit(core.Model):
            id = core.SlugAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

        class Quantity(core.Model):
            value = core.FloatAttribute()
            unit = core.ManyToOneAttribute(Unit, related_name='quantities_1')
            units = core.ManyToManyAttribute(Unit, related_name='quantities_2')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.multiple_cells
                attribute_order = ('value', 'unit', 'units')

            def serialize(self):
                return '{} {}'.format(self.value, self.unit.id)

        class Node(core.Model):
            id = core.SlugAttribute()
            quantity_1 = core.OneToOneAttribute(Quantity, related_name='node_1')
            quantity_2 = core.ManyToOneAttribute(Quantity, related_name='node_2_list')
            comments = core.LongStringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'quantity_1', 'quantity_2', 'comments')

        u1 = Unit(id='m')
        u2 = Unit(id='s')
        u3 = Unit(id='g')
        u4 = Unit(id='l')
        u5 = Unit(id='k')
        q1 = Quantity(value=1., unit=u1, units=[u3, u4])
        q2 = Quantity(value=2., unit=u1, units=[u4])
        q3 = Quantity(value=2., unit=u2, units=[])
        q4 = Quantity(value=3., unit=u2, units=[u5])

        nodes = [
            Node(id='node0', quantity_1=q1, quantity_2=q3, comments='node 0'),
            Node(id='node1', quantity_1=q2, quantity_2=q3, comments='node 1'),
            Node(id='node2', quantity_1=None, quantity_2=q4, comments='node 2'),
            Node(id='node3', quantity_1=None, quantity_2=None, comments='node 3'),
        ]

        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        writer = WorkbookWriter()
        writer.run(filename, nodes, models=[Node, Unit])
        nodes_2 = WorkbookReader().run(filename, models=[Node, Unit])[Node]
        for node, node_2 in zip(nodes, nodes_2):
            self.assertTrue(node_2.is_equal(node))

        # error
        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        wb = read_workbook(filename)
        wb['!Nodes'][4][1] = 'a'
        filename = os.path.join(self.tmp_dirname, 'test3.xlsx')
        write_workbook(filename, wb)
        with self.assertRaisesRegex(ValueError, 'model cannot be loaded'):
            WorkbookReader().run(filename, models=[Node, Unit])

        # column orientation
        Node.Meta.table_format = core.TableFormat.column
        filename = os.path.join(self.tmp_dirname, 'test2.xlsx')
        writer = WorkbookWriter()
        writer.run(filename, nodes, models=[Node, Unit])
        nodes_2 = WorkbookReader().run(filename, models=[Node, Unit])[Node]
        for node, node_2 in zip(nodes, nodes_2):
            self.assertTrue(node_2.is_equal(node))

    def test_toc(self):
        class Model1(core.Model):
            id = core.SlugAttribute()

        class Model2(core.Model):
            id = core.SlugAttribute()

        objs = [
            Model1(id='model1_0'),
            Model1(id='model1_1'),
            Model2(id='model2_0'),
            Model2(id='model2_1'),
        ]

        path = os.path.join(self.tmp_dirname, 'test.xlsx')
        obj_tables.io.Writer().run(path, objs, models=[Model1, Model2])
        objs_2 = obj_tables.io.Reader().run(path, models=[Model1, Model2])
        for obj, obj_2 in zip(objs, objs_2):
            self.assertTrue(obj_2.is_equal(obj))

        path = os.path.join(self.tmp_dirname, 'test*.csv')
        obj_tables.io.Writer().run(path, objs, models=[Model1, Model2])
        objs_2 = obj_tables.io.Reader().run(path, models=[Model1, Model2])
        for obj, obj_2 in zip(objs, objs_2):
            self.assertTrue(obj_2.is_equal(obj))

    def test_ignore_empty_rows(self):
        class Node(core.Model):
            id = core.SlugAttribute()
            name = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'name', )

        filename = os.path.join(self.tmp_dirname, 'test.xlsx')
        wb = Workbook()
        ws = wb['!Nodes'] = Worksheet()
        ws.append(Row(["!!ObjTables TableType='Data' ModelId='Node'"]))
        ws.append(Row(['!Id', '!Name']))
        ws.append(Row(['a', 'A']))
        ws.append(Row(['b', 'B']))
        ws.append(Row(['', '']))
        ws.append(Row(['d', 'D']))
        write_workbook(filename, wb)

        objs = obj_tables.io.Reader().run(filename, models=[Node])
        self.assertEqual(len(objs), 3)

        with self.assertRaisesRegex(ValueError, r'contains error\(s\)'):
            obj_tables.io.Reader().run(filename, models=[Node], ignore_empty_rows=False)

    def test_model_metadata(self):
        class Node(core.Model):
            id = core.SlugAttribute()
            name = core.StringAttribute()

        filename = os.path.join(self.tmp_dirname, 'test.xlsx')

        objs = [
            Node(id='a', name='A'),
            Node(id='b', name='B'),
            Node(id='c', name='C'),
        ]
        model_metadata = {Node: {
            'attr1': 'val1',
            'attr2': 'val2',
        }}

        writer = obj_tables.io.Writer()
        writer.run(filename, objs, model_metadata=model_metadata, models=[Node])

        reader = obj_tables.io.Reader()
        objs2 = reader.run(filename, models=[Node])
        for obj, obj2 in zip(objs, objs2):
            self.assertTrue(obj2.is_equal(obj))

        model_metadata = {Node: {
            'TableType': 'Data',
            'ModelId': 'Node',
            'ModelName': 'Nodes',
            'ObjTablesVersion': obj_tables.__version__,
            'attr1': 'val1',
            'attr2': 'val2',
        }}
        reader._model_metadata[Node].pop('Date')
        self.assertEqual(reader._model_metadata, model_metadata)


class TestMetadataModels(unittest.TestCase):

    class Model1(core.Model):
        id = core.SlugAttribute()

    def setUp(self):
        self.objs = [
            self.Model1(id='model1_0'),
            self.Model1(id='model1_1'),
        ]

        self.tmp_dirname = tempfile.mkdtemp()

        # prepare test data repo
        self.github_test_data_repo = GitHubRepoForTests('test_data_repo')
        self.test_data_repo_dir = os.path.join(self.tmp_dirname, 'test_data_repo')
        os.mkdir(self.test_data_repo_dir)
        test_data_repo = self.github_test_data_repo.make_test_repo(self.test_data_repo_dir)

        # prepare test schema repo
        test_schema_repo_url = 'https://github.com/KarrLab/test_repo'
        self.test_schema_repo_dir = os.path.join(self.tmp_dirname, 'test_schema_repo')
        test_schema_repo = git.Repo.clone_from(test_schema_repo_url, self.test_schema_repo_dir)

        # put schema dir on sys.path
        sys.path.append(self.test_schema_repo_dir)

    def tearDown(self):
        shutil.rmtree(self.tmp_dirname)

        self.github_test_data_repo.delete_test_repo()

        # remove self.test_schema_repo_dir from sys.path
        for idx in range(len(sys.path)-1, -1, -1):
            if sys.path[idx] == self.test_schema_repo_dir:
                del sys.path[idx]

    def test_workbook_writer_make_metadata_objects(self):
        writer = obj_tables.io.Writer()
        reader = obj_tables.io.Reader()

        ### test metadata return ###
        # read data repo metadata
        file_in_repo = os.path.join(self.test_data_repo_dir, 'test.xlsx')
        writer.run(file_in_repo, self.objs, models=[self.Model1], data_repo_metadata=True)
        objs_read = reader.run(file_in_repo, models=[utils.DataRepoMetadata, self.Model1])
        data_repo_metadata = objs_read[0]
        self.assertTrue(data_repo_metadata.url.startswith('https://github.com/'))
        self.assertEqual(data_repo_metadata.branch, 'master')
        self.assertTrue(isinstance(data_repo_metadata.revision, str))
        self.assertEqual(len(data_repo_metadata.revision), 40)
        for obj, obj_read in zip(self.objs, objs_read[1:]):
            self.assertTrue(obj_read.is_equal(obj))

        models_expected = [utils.DataRepoMetadata, utils.SchemaRepoMetadata, self.Model1]
        objs_read = reader.run(file_in_repo, models=models_expected,
                               ignore_missing_models=True)
        obj_types = [o.__class__ for o in objs_read]
        self.assertTrue(utils.SchemaRepoMetadata not in obj_types)

        # write data and schema repo metadata to data file
        writer.run(file_in_repo, self.objs, models=[self.Model1], data_repo_metadata=True,
                   schema_package='test_repo')

        # read data and schema metadata
        objs_read = reader.run(file_in_repo, models=models_expected,)
        for obj, model in zip(objs_read, models_expected[0:2]):
            self.assertTrue(isinstance(obj, model))
            self.assertTrue(obj.url.startswith('https://github.com/'))
            self.assertEqual(obj.branch, 'master')
            self.assertTrue(isinstance(obj.revision, str))
            self.assertEqual(len(obj.revision), 40)

        # test csv files with metadata
        csv_path = os.path.join(self.test_data_repo_dir, 'test*.csv')
        writer.run(csv_path, self.objs, models=[self.Model1], data_repo_metadata=True,
                   schema_package='test_repo')
        objs_read = reader.run(csv_path, models=models_expected)
        for obj, model in zip(objs_read, models_expected):
            self.assertTrue(isinstance(obj, model))

        ### test warnings & errors ###
        ## data repo ##
        # data repo metadata not written: data file not in a repo
        with pytest.warns(obj_tables.io.IoWarning) as w:
            file_not_in_repo = os.path.join(self.tmp_dirname, 'test.xlsx')
            writer.run(file_not_in_repo, self.objs, models=[self.Model1], data_repo_metadata=True)
            self.assertEqual(len(w), 1)
            warning_msg = str(w[0].message)
            self.assertIn('Cannot obtain git repo metadata for data repo', warning_msg)
            self.assertIn('is not in a Git repository', warning_msg)

        # data repo contains changes
        # write other file in data repo
        other_file = os.path.join(self.test_data_repo_dir, 'foo.txt')
        with open(other_file, 'w') as f:
            f.write('hello world!')
        with pytest.warns(obj_tables.io.IoWarning) as w:
            # write data file in test data repo
            writer.run(file_in_repo, self.objs, models=[self.Model1], data_repo_metadata=True)
            self.assertEqual(len(w), 1)
            warning_msg = str(w[0].message)
            self.assertIn("Git repo metadata for data repo was obtained", warning_msg)
            self.assertIn("Ensure that the data file", warning_msg)

        ## schema repo ##
        with pytest.warns(obj_tables.io.IoWarning) as w:
            # test schema package not found
            writer.run(file_in_repo, self.objs, models=[self.Model1], data_repo_metadata=False,
                       schema_package='not a schema package')
            self.assertEqual(len(w), 1)
            warning_msg = str(w[0].message)
            self.assertRegex(warning_msg, "package '.+' not found")
            self.assertIn("Cannot obtain git repo metadata for schema repo", warning_msg)

        with pytest.warns(obj_tables.io.IoWarning) as w:
            # test schema repo modified by file_in_schema_repo
            file_in_schema_repo = os.path.join(self.test_schema_repo_dir, 'file_in_schema_repo.txt')
            with open(file_in_schema_repo, 'w') as f:
                f.write('hello world!')
            package_name = 'test_repo'
            writer.run(file_in_repo, self.objs, models=[self.Model1], data_repo_metadata=False,
                       schema_package=package_name)
            self.assertEqual(len(w), 1)
            warning_msg = str(w[0].message)
            self.assertRegex(warning_msg,
                             "Cannot obtain git repo metadata for schema repo '.+' used by data file:")
            self.assertIn("Cannot gather metadata for schema repo from Git repo containing",
                          warning_msg)

        # clean up
        os.remove(file_in_repo)
        os.remove(other_file)

    def test_drop_metadata_model(self):
        file_with_metadata = os.path.join(os.path.dirname(__file__), 'fixtures', 'metadata',
                                          'both-metadata.xlsx')
        objs_read = obj_tables.io.Reader().run(file_with_metadata, models=[
            utils.DataRepoMetadata, utils.SchemaRepoMetadata, self.Model1],
            ignore_attribute_order=True, group_objects_by_model=True)
        self.assertEqual(len(objs_read[utils.DataRepoMetadata]), 1)
        self.assertTrue(isinstance(objs_read[utils.DataRepoMetadata][0], utils.DataRepoMetadata))

    def test_json_writer_make_metadata_objects(self):

        # write data repo metadata in obj_tables file
        path_1 = os.path.join(self.test_data_repo_dir, 'out.json')
        obj_tables.io.JsonWriter().run(path_1, self.objs, models=[self.Model1], data_repo_metadata=True)
        objs_read = obj_tables.io.JsonReader().run(path_1, models=[utils.DataRepoMetadata, self.Model1])
        data_repo_metadata = objs_read[0]
        self.assertTrue(data_repo_metadata.url.startswith('https://github.com/'))
        self.assertEqual(data_repo_metadata.branch, 'master')
        self.assertTrue(isinstance(data_repo_metadata.revision, str))
        self.assertEqual(len(data_repo_metadata.revision), 40)
        for obj, obj_read in zip(self.objs, objs_read[1:]):
            self.assertTrue(obj_read.is_equal(obj))

        # test data and schema repo metadata in obj_tables file
        obj_tables.io.JsonWriter().run(path_1, self.objs, models=[self.Model1], data_repo_metadata=True,
                                       schema_package='test_repo')
        metadata_models_expected = [utils.DataRepoMetadata, utils.SchemaRepoMetadata]
        objs_read = obj_tables.io.JsonReader().run(path_1, models=metadata_models_expected + [self.Model1])
        for obj, model in zip(objs_read, metadata_models_expected):
            self.assertTrue(isinstance(obj, model))
            self.assertTrue(obj.url.startswith('https://github.com/'))
            self.assertEqual(obj.branch, 'master')
            self.assertTrue(isinstance(obj.revision, str))
            self.assertEqual(len(obj.revision), 40)

    def test_json_read_extra_worksheets(self):
        class Parent(core.Model):
            id = core.SlugAttribute()

        class Child(core.Model):
            id = core.SlugAttribute()
            parent = core.ManyToOneAttribute(Parent, related_name='children')

        class UnsupportedType(core.Model):
            field = core.SlugAttribute()

        p = Parent(id='p')
        c_1 = p.children.create(id='c_1')
        c_2 = p.children.create(id='c_2')
        objs = [p, c_1, c_2]

        # list of objects
        path = os.path.join(self.tmp_dirname, 'out.json')
        obj_tables.io.JsonWriter().run(path, objs, models=[Parent, Child])

        objs = obj_tables.io.JsonReader().run(path, models=[Parent, Child])
        p_b = next(obj for obj in objs if isinstance(obj, Parent))
        self.assertTrue(p_b.is_equal(p))

        with open(path, 'rb') as file:
            objs = json.load(file)
        objs.append({'__type': 'UnsupportedType', '__id': 3, 'field': 'data'})
        with open(path, 'w') as file:
            json.dump(objs, file)

        with self.assertRaisesRegex(ValueError, 'Unsupported type'):
            obj_tables.io.JsonReader().run(path, models=[Parent, Child])

        # single object
        path = os.path.join(self.tmp_dirname, 'out.json')
        obj_tables.io.JsonWriter().run(path, p, models=[Parent, Child])

        p_b = obj_tables.io.JsonReader().run(path, models=[Parent, Child])
        self.assertTrue(p_b.is_equal(p))

        obj = {'__type': 'UnsupportedType', '__id': 0, 'field': 'data'}
        with open(path, 'w') as file:
            json.dump(obj, file)

        with self.assertRaisesRegex(ValueError, 'Unsupported type'):
            obj_tables.io.JsonReader().run(path, models=[Parent, Child])

        obj = obj_tables.io.JsonReader().run(path, models=[Parent, Child, UnsupportedType])
        self.assertNotEqual(obj, None)
        self.assertEqual(obj.field, 'data')


class TestMisc(unittest.TestCase):

    def setUp(self):
        self.dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.dirname)

    def test_read_write_row_oriented(self):
        class Parent1(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')
            children = core.OneToManyAttribute('Child1', verbose_name='children', related_name='parent')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'children')

        class Child1(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

        parents = [
            Parent1(id='parent_0'),
            Parent1(id='parent_1'),
            Parent1(id='parent_2'),
        ]
        parents[0].children.create(id='child_0_0')
        parents[0].children.create(id='child_0_1')
        parents[0].children.create(id='child_0_2')
        parents[1].children.create(id='child_1_0')
        parents[1].children.create(id='child_1_1')
        parents[1].children.create(id='child_1_2')
        parents[2].children.create(id='child_2_0')
        parents[2].children.create(id='child_2_1')
        parents[2].children.create(id='child_2_2')

        filename = os.path.join(self.dirname, 'test.xlsx')

        writer = WorkbookWriter()
        writer.run(filename, parents, models=[Parent1, Child1])

        objects = WorkbookReader().run(filename, models=[Parent1, Child1])
        objects[Parent1].sort(key=lambda parent: parent.id)
        for orig_parent, copy_parent in zip(parents, objects[Parent1]):
            self.assertTrue(orig_parent.is_equal(copy_parent))

    def test_warn_about_errors(self):
        class Node2(core.Model):
            id = core.StringAttribute(max_length=1, primary=True, unique=True, verbose_name='Identifier')

        node = Node2(id='id')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer = WorkbookWriter()

        with pytest.warns(IoWarning):
            writer.run(filename, [node], models=[Node2])

    def test_writer_error_if_not_serializable(self):
        class ChildrenAttribute3(core.OneToManyAttribute):
            pass

        class Parent3(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')
            children = ChildrenAttribute3('Child3', verbose_name='children', related_name='parent')

        class Child3(core.Model):
            id = core.StringAttribute(verbose_name='Identifier')

        parent = Parent3(id='parent')
        parent.children.create(id='child_1')
        parent.children.create(id='child_2')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer = WorkbookWriter()

        with self.assertRaisesRegex(ValueError, 'cannot be serialized'):
            writer.run(filename, [parent], models=[Parent3, Child3])

    def test_reader_error_if_not_serializable(self):
        class ChildrenAttribute(core.OneToManyAttribute):
            def serialize(self, value, encoded=None):
                return super(ChildrenAttribute, self).serialize(value)

            def deserialize(self, values, objects, decoded=None):
                return super(ChildrenAttribute, self).deserialize(value, objects)

        class Child(core.Model):
            id = core.StringAttribute(primary=True, verbose_name='Identifier')

            class Meta(core.Model.Meta):
                verbose_name_plural = 'Children'
                attribute_order = ('id',)

        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')
            children = ChildrenAttribute(Child, verbose_name='children', related_name='parent')

            class Meta(core.Model.Meta):
                verbose_name_plural = 'Parents'
                attribute_order = ('id', 'children')

        parent = Parent(id='parent')
        parent.children.create(id='child_1')
        parent.children.create(id='child_2')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer = WorkbookWriter()
        writer.run(filename, [parent], models=[Parent, Child])

        class ChildrenAttribute(core.OneToManyAttribute):
            pass

        class Child(core.Model):
            id = core.StringAttribute(verbose_name='Identifier')

            class Meta(core.Model.Meta):
                verbose_name_plural = 'Children'
                attribute_order = ('id',)

        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')
            children = ChildrenAttribute(Child, verbose_name='children', related_name='parent')

            class Meta(core.Model.Meta):
                verbose_name_plural = 'Parents'
                attribute_order = ('id', 'children')

        with self.assertRaisesRegex(ValueError, 'cannot be serialized'):
            WorkbookReader().run(filename, models=[Parent, Child],
                                 ignore_missing_models=True)

    def test_read_missing_sheet(self):
        class Node6(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')

        class Node7(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')

        nodes = [
            Node6(id='node_6_0'),
            Node6(id='node_6_1'),
            Node6(id='node_6_2'),
        ]

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer = WorkbookWriter()
        writer.run(filename, nodes, models=[Node6])

        objects = WorkbookReader().run(filename, models=[Node6, Node7], ignore_missing_models=True)
        objects[Node6].sort(key=lambda node: node.id)
        for orig_node, copy_node in zip(nodes, objects[Node6]):
            self.assertTrue(orig_node.is_equal(copy_node))
        self.assertEqual(objects[Node7], [])

    def test_ambiguous_column_headers(self):
        class Node8(core.Model):
            id1 = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')
            id2 = core.StringAttribute(primary=False, unique=True, verbose_name='Identifier')
            id3 = core.StringAttribute(primary=False, unique=True)

        nodes = [
            Node8(id1='node_0_1', id2='node_0_2'),
            Node8(id1='node_1_1', id2='node_1_2'),
        ]

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer = WorkbookWriter()

        with pytest.warns(IoWarning):
            writer.run(filename, nodes, models=[Node8])

        with self.assertRaisesRegex(ValueError, 'Duplicate, case insensitive, headers:'):
            WorkbookReader().run(filename, models=[Node8])

    def test_row_and_column_headings(self):
        class TestModel(core.Model):
            column_B = core.StringAttribute()
            column_C = core.StringAttribute()

            class Meta(core.Model.Meta):
                verbose_name_plural = 'Sheet'

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer = WorkbookWriter()
        xslx_writer = get_writer('.xlsx')(filename)
        xslx_writer.initialize_workbook()
        writer.write_sheet(xslx_writer,
                           TestModel,
                           data=[['Cell_2_B', 'Cell_2_C'], ['Cell_3_B', 'Cell_3_C']],
                           headings=[["!!ObjTables TableType='Data' ModelId='TestModel'"],
                                     ['!Column_B', '!Column_C']],
                           metadata_headings=[],
                           validation=None)
        xslx_writer.finalize_workbook()

        xlsx_reader = get_reader('.xlsx')(filename)
        workbook = xlsx_reader.run()
        self.assertEqual(list(workbook['!Sheet'][0]), ["!!ObjTables TableType='Data' ModelId='TestModel'", None])
        self.assertEqual(list(workbook['!Sheet'][1]), ['!Column_B', '!Column_C'])
        self.assertEqual(list(workbook['!Sheet'][2]), ['Cell_2_B', 'Cell_2_C'])
        self.assertEqual(list(workbook['!Sheet'][3]), ['Cell_3_B', 'Cell_3_C'])

        reader = WorkbookReader()
        xlsx_reader = get_reader('.xlsx')(filename)
        xlsx_reader.initialize_workbook()
        reader._model_metadata = {}
        data, row_headings, column_headings, _ = reader.read_sheet(TestModel, xlsx_reader, '!Sheet',
                                                                   num_row_heading_columns=0,
                                                                   num_column_heading_rows=1)
        self.assertEqual(len(data), 2)
        self.assertEqual(list(data[0]), ['Cell_2_B', 'Cell_2_C'])
        self.assertEqual(list(data[1]), ['Cell_3_B', 'Cell_3_C'])
        self.assertEqual(row_headings, [])
        self.assertEqual(len(column_headings), 1)
        self.assertEqual(list(column_headings[0]), ['!Column_B', '!Column_C'])

    def test_get_model_sheet_name_error(self):
        class Node9(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Identifier')

            class Meta(core.Model.Meta):
                verbose_name_plural = 'Nodes'

        with self.assertRaisesRegex(ValueError, 'matches multiple sheets'):
            WorkbookReader.get_model_sheet_name(['!Nodes', '!nodes'], Node9)

    def test_unclean_data(self):
        workbook = Workbook()
        workbook['!Node10'] = worksheet = Worksheet()
        workbook['!Node10'].append(Row(["!!ObjTables TableType='Data' ModelId='Node10'"]))
        worksheet.append(Row(['!Id', '!Value']))
        worksheet.append(Row(['A', '1.0']))
        worksheet.append(Row(['B', 'x']))

        filename = os.path.join(self.dirname, 'test.xlsx')
        xslx_writer = get_writer('.xlsx')(filename)
        xslx_writer.run(workbook, style={
            'Node10': WorksheetStyle(extra_rows=0, extra_columns=0),
        })

        class Node10(core.Model):
            id = core.StringAttribute(primary=True, unique=True, verbose_name='Id')
            value = core.FloatAttribute(verbose_name='Value')

        with self.assertRaisesRegex(ValueError, 'Value must be a `float`'):
            WorkbookReader().run(filename, models=[Node10])

    def test_write_read_subset_of_attributes(self):
        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id', )

        class Child(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            parent = core.ManyToOneAttribute(Parent, related_name='children')

            class Meta(core.Model.Meta):
                attribute_order = ('id', )

        class GrandChild(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            parent = core.ManyToOneAttribute(Child, related_name='children')

            class Meta(core.Model.Meta):
                attribute_order = ('id', )

        filename = os.path.join(self.dirname, 'test.xlsx')

        ######
        p = Parent(id='p')

        WorkbookWriter().run(filename, p, models=Parent, include_all_attributes=True)
        objs2 = WorkbookReader().run(filename, models=Parent, include_all_attributes=True)
        self.assertEqual(len(objs2[Parent]), 1)
        self.assertNotIn(Child, objs2)
        self.assertNotIn(GrandChild, objs2)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

        WorkbookWriter().run(filename, p, models=None, include_all_attributes=True)
        objs2 = WorkbookReader().run(filename, models=Parent, include_all_attributes=True)
        self.assertEqual(len(objs2[Parent]), 1)
        self.assertNotIn(Child, objs2)
        self.assertNotIn(GrandChild, objs2)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

        ######
        p = Parent(id='p')
        c0 = p.children.create(id='c0')
        c1 = p.children.create(id='c1')
        g00 = c0.children.create(id='g00')
        g01 = c0.children.create(id='g01')
        g10 = c0.children.create(id='g10')
        g11 = c0.children.create(id='g11')

        WorkbookWriter().run(filename, p, models=[Parent, Child, GrandChild], include_all_attributes=True)
        objs2 = WorkbookReader().run(filename, models=[Parent, Child, GrandChild], include_all_attributes=True)
        self.assertEqual(len(objs2[Parent]), 1)
        self.assertEqual(len(objs2[Child]), 2)
        self.assertEqual(len(objs2[GrandChild]), 4)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

        WorkbookWriter().run(filename, p, models=[Parent, Child, GrandChild], include_all_attributes=False)
        objs2 = WorkbookReader().run(filename, models=[Parent, Child, GrandChild], include_all_attributes=False)
        self.assertEqual(len(objs2[Parent]), 1)
        self.assertEqual(len(objs2[Child]), 2)
        self.assertEqual(len(objs2[GrandChild]), 4)
        p2 = objs2[Parent][0]
        self.assertFalse(p.is_equal(p2))
        self.assertEqual(p2.children, [])
        self.assertEqual(set(c.id for c in objs2[Child]), set(['c0', 'c1']))
        for c in objs2[Child]:
            self.assertEqual(c.parent, None)
            self.assertEqual(c.children, [])
        self.assertEqual(set(g.id for g in objs2[GrandChild]), set(['g00', 'g01', 'g10', 'g11']))
        for g in objs2[GrandChild]:
            self.assertEqual(g.parent, None)

        ######
        with self.assertRaisesRegex(ValueError, 'At least one `Model` must be provided'):
            WorkbookWriter().run(filename, None, models=None)

        with self.assertRaisesRegex(ValueError, 'No matching models'):
            WorkbookReader().run(filename, models=None)


class ReadEmptyCellTestCase(unittest.TestCase):
    def setUp(self):
        self.dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.dirname)

    def test_clean_enum(self):
        class TestEnum(enum.Enum):
            val = 0

        attr = core.EnumAttribute(TestEnum)
        self.assertNotEqual(attr.clean({})[1], None)

    def test_get_default_cleaned_value(self):
        class ConcreteAttribute(core.Attribute):
            def deserialize(self, value):
                pass

            def serialize(self):
                pass

            def validate(self):
                pass

            def to_builtin(self, encoded=None):
                pass

            def from_builtin(self, json, decoded=None):
                pass

            def merge(self, other, validate=True):
                pass

            def copy_value(self, value, objects_and_copies):
                pass

        attr = ConcreteAttribute(default_cleaned_value=lambda: 1.5)
        self.assertEqual(attr.get_default_cleaned_value(), 1.5)

    def test_read_empty_float(self):
        class TestModel(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            value_1 = core.FloatAttribute(default_cleaned_value=float('nan'))
            value_2 = core.FloatAttribute(default_cleaned_value=2.)

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'value_1', 'value_2')

        workbook = Workbook()
        workbook['!Test models'] = worksheet = Worksheet()
        worksheet.append(Row(["!!ObjTables TableType='Data' ModelId='TestModel'"]))
        worksheet.append(Row(['!Id', '!Value 1', '!Value 2']))
        worksheet.append(Row(['A', None, None]))
        worksheet.append(Row(['B', 1., 3.]))
        worksheet.append(Row(['C', None, None]))

        filename = os.path.join(self.dirname, 'test.xlsx')
        xslx_writer = get_writer('.xlsx')(filename)
        xslx_writer.run(workbook, style={
            TestModel.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })

        objects = WorkbookReader().run(filename, models=[TestModel])[TestModel]
        objects.sort(key=lambda m: m.id)

        self.assertTrue(math.isnan(objects[0].value_1))
        self.assertEqual(objects[1].value_1, 1.)
        self.assertTrue(math.isnan(objects[2].value_1))

        self.assertEqual(objects[0].value_2, 2.)
        self.assertEqual(objects[1].value_2, 3.)
        self.assertEqual(objects[2].value_2, 2.)


class InheritedIoTestCase(unittest.TestCase):
    def setUp(self):
        self.dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.dirname)

    def test_OneToOneAttribute(self):
        class B(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

        class BB(B):
            pass

        class A(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            b = core.OneToOneAttribute(B, related_name='a')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'b')

        class AA(A):
            pass

        aa1 = AA(id='aa1')
        bb1 = BB(id='bb')
        aa1.b = bb1

        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [aa1], models=[AA, BB])

        aa2 = WorkbookReader().run(filename, models=[AA, BB])[AA][0]
        self.assertTrue(aa2.is_equal(aa1))
        self.assertEqual(aa1.b.a, aa1)

    def test_ManyToOneAttribute(self):
        class B(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

        class BB(B):
            pass

        class A(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            b = core.ManyToOneAttribute(B, related_name='a_s')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'b')

        class AA(A):
            pass

        aa1 = AA(id='aa1')
        bb1 = BB(id='bb')
        aa1.b = bb1

        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [aa1], models=[AA, BB])

        aa2 = WorkbookReader().run(filename, models=[AA, BB])[AA][0]
        self.assertTrue(aa2.is_equal(aa1))
        self.assertIn(aa1, aa1.b.a_s)

    def test_OneToManyAttribute(self):
        class B(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

        class BB(B):
            pass

        class A(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            bs = core.OneToManyAttribute(B, related_name='a')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'bs')

        class AA(A):
            pass

        aa1 = AA(id='aa1')
        bb1 = BB(id='bb')
        aa1.bs.append(bb1)

        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [aa1], models=[AA, BB])

        aa2 = WorkbookReader().run(filename, models=[AA, BB])[AA][0]
        self.assertTrue(aa2.is_equal(aa1))
        self.assertEqual(bb1.a, aa1)

    def test_ManyToManyAttribute(self):
        class B(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

        class BB(B):
            pass

        class A(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            bs = core.ManyToManyAttribute(B, related_name='a_s')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'bs')

        class AA(A):
            pass

        aa1 = AA(id='aa1')
        bb1 = BB(id='bb')
        aa1.bs.append(bb1)

        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [aa1], models=[AA, BB])

        aa2 = WorkbookReader().run(filename, models=[AA, BB])[AA][0]
        self.assertTrue(aa2.is_equal(aa1))
        self.assertIn(aa1, bb1.a_s)


class StrictReadingTestCase(unittest.TestCase):
    def setUp(self):
        self.dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.dirname)

    def test_missing_sheet(self):
        class Model1(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

        class Model2(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
        m1 = Model1(id='m1')
        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [m1], models=[Model1])

        result = WorkbookReader().run(filename, models=[Model1])
        self.assertEqual(set(result.keys()), set([Model1]))
        self.assertEqual(len(result[Model1]), 1)
        self.assertTrue(m1.is_equal(result[Model1][0]))

        with self.assertRaises(ValueError):
            WorkbookReader().run(filename, models=[Model1, Model2])

        result = WorkbookReader().run(filename, models=[Model1, Model2], ignore_missing_models=True)
        self.assertEqual(set(result.keys()), set([Model1, Model2]))
        self.assertEqual(len(result[Model1]), 1)
        self.assertEqual(len(result[Model2]), 0)
        self.assertTrue(m1.is_equal(result[Model1][0]))

    def test_extra_sheet(self):
        class Model1(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

        class Model2(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
        m1 = Model1(id='m1')
        m2 = Model2(id='m2')
        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [m1, m2], models=[Model1, Model2])

        result = WorkbookReader().run(filename, models=[Model1, Model2])
        self.assertEqual(set(result.keys()), set([Model1, Model2]))
        self.assertEqual(len(result[Model1]), 1)
        self.assertEqual(len(result[Model2]), 1)
        self.assertTrue(m1.is_equal(result[Model1][0]))
        self.assertTrue(m2.is_equal(result[Model2][0]))

        with self.assertRaises(ValueError):
            WorkbookReader().run(filename, models=[Model1])

    def test_different_sheet_order(self):
        class Model1(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

        class Model2(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

        class Model3(core.Model):
            id = core.StringAttribute(primary=True, unique=True)

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.column

        m1 = Model1(id='m1')
        m2 = Model2(id='m2')
        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, [m1, m2], models=[Model1, Model2])

        WorkbookReader().run(filename, models=[Model1, Model2])
        with self.assertRaises(ValueError):
            WorkbookReader().run(filename, models=[Model2, Model1])
        WorkbookReader().run(filename, models=[Model2, Model1], ignore_sheet_order=True)

        WorkbookWriter().run(filename, [m1], models=[Model1])
        WorkbookReader().run(filename, models=[Model1, Model3], ignore_missing_models=True)

    def test_no_header_rows(self):
        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        ws.append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', '!Attr']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        result = WorkbookReader().run(filename, models=[Model])
        self.assertEqual(result, {Model: []})

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        reader = get_reader('.xlsx')(filename)
        reader.initialize_workbook()
        wb_reader = WorkbookReader()
        wb_reader._model_metadata = {}
        with self.assertRaisesRegex(ValueError, r'must have 1 header row\(s\)'):
            wb_reader.read_sheet(Model, reader, '!Models', num_column_heading_rows=1)

    def test_no_header_cols(self):
        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr')
                table_format = core.TableFormat.column

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id']))
        ws.append(Row(['!Attr']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(ValueError, r'must have 1 header column\(s\)'):
            WorkbookReader().run(filename, models=[Model])

    def test_missing_attribute(self):
        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr1 = core.StringAttribute()
            attr2 = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr1', 'attr2')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', '!Attr1', '!Attr2']))
        ws.append(Row(['m1', '1', '2']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', '!Attr2']))
        ws.append(Row(['m1', '2']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaises(ValueError):
            WorkbookReader().run(filename, models=[Model])
        WorkbookReader().run(filename, models=[Model], ignore_missing_attributes=True)

    def test_missing_attribute_inline(self):
        class Quantity(core.Model):
            value = core.FloatAttribute()
            units = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('value', 'units')
                table_format = core.TableFormat.multiple_cells

            def serialize(self):
                return '{} {}'.format(self.value, self.units)

        class Quantity2(core.Model):
            value = core.FloatAttribute()
            units = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('value', 'units')
                table_format = core.TableFormat.multiple_cells

            def serialize(self):
                return '{} {}'.format(self.value, self.units)

        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr1 = core.StringAttribute()
            attr2 = core.StringAttribute()
            quantity = core.OneToOneAttribute(Quantity, related_name='model')
            quantity2 = core.OneToOneAttribute(Quantity2, related_name='model')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr1', 'attr2')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row([None, None, None, '!Quantity', '!Quantity', '!Quantity2', '!Quantity2']))
        ws.append(Row(['!Id', '!Attr1', '!Attr2', '!Value', '!Units', '!Value', '!Units']))
        ws.append(Row(['m1', '1', '2', 1.2, 'g', 1.2, 'g']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row([None, None, '!Quantity2', '!Quantity2']))
        ws.append(Row(['!Id', '!Attr2', '!Value', '!Units']))
        ws.append(Row(['m1', '2', 1.2, 'g']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaises(ValueError):
            WorkbookReader().run(filename, models=[Model])
        WorkbookReader().run(filename, models=[Model], ignore_missing_attributes=True, ignore_empty_rows=False)

    def test_extra_attribute(self):
        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr1 = core.StringAttribute()
            attr2 = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr1', 'attr2')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', '!Attr1', '!Attr2']))
        ws.append(Row(['m1', '1', '2']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', '!Attr1', '!Attr2', '!Attr3']))
        ws.append(Row(['m1', '1', '2', '3']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaises(ValueError):
            WorkbookReader().run(filename, models=[Model])
        WorkbookReader().run(filename, models=[Model], ignore_extra_attributes=True)

    def test_different_attribute_order(self):
        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr1 = core.StringAttribute()
            attr2 = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr1', 'attr2')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', '!Attr1', '!Attr2']))
        ws.append(Row(['m1', '1', '2']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', '!Attr2', '!Attr1']))
        ws.append(Row(['m1', '2', '1']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(ValueError, (
            "The columns of worksheet '!Models' must be defined in this order:"
            "\n      A1: !Id"
            "\n      B1: !Attr1"
            "\n      C1: !Attr2"
        )):
            WorkbookReader().run(filename, models=[Model])
        WorkbookReader().run(filename, models=[Model], ignore_attribute_order=True)

    def test_different_attribute_order_transpose(self):
        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr1 = core.StringAttribute()
            attr2 = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr1', 'attr2')
                table_format = core.TableFormat.column

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', 'm1']))
        ws.append(Row(['!Attr1', '1']))
        ws.append(Row(['!Attr2', '2']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row(['!Id', 'm1']))
        ws.append(Row(['!Attr2', '2']))
        ws.append(Row(['!Attr1', '1']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(ValueError, (
            "The rows of worksheet '!Models' must be defined in this order:"
            "\n      A1: !Id"
            "\n      A2: !Attr1"
            "\n      A3: !Attr2"
        )):
            WorkbookReader().run(filename, models=[Model])
        WorkbookReader().run(filename, models=[Model], ignore_attribute_order=True)

    def test_different_attribute_order_inline(self):
        class Quantity(core.Model):
            value = core.FloatAttribute()
            units = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('value', 'units')
                table_format = core.TableFormat.multiple_cells

            def serialize(self):
                return '{} {}'.format(self.value, self.units)

        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr1 = core.StringAttribute()
            attr2 = core.StringAttribute()
            quantity = core.OneToOneAttribute(Quantity, related_name='model')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr1', 'attr2')

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row([None, None, None, '!Quantity', '!Quantity']))
        ws.append(Row(['!Id', '!Attr1', '!Attr2', '!Value', '!Units']))
        ws.append(Row(['m1', '1', '2', 1.1, 's']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row([None, None, None, '!Quantity', '!Quantity']))
        ws.append(Row(['!Id', '!Attr2', '!Attr1', '!Value', '!Units']))
        ws.append(Row(['m1', '2', '1', 1.1, 's']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(ValueError, (
            "The columns of worksheet '!Models' must be defined in this order:"
            "\n      A1: "
            "\n      A2: !Id"
            "\n      B1: "
            "\n      B2: !Attr1"
            "\n      C1: "
            "\n      C2: !Attr2"
            "\n      D1: !Quantity"
            "\n      D2: !Value"
            "\n      E1: !Quantity"
            "\n      E2: !Units"
        )):
            WorkbookReader().run(filename, models=[Model])
        WorkbookReader().run(filename, models=[Model], ignore_attribute_order=True)

    def test_different_attribute_order_inline_transpose(self):
        class Quantity(core.Model):
            value = core.FloatAttribute()
            units = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('value', 'units')
                table_format = core.TableFormat.multiple_cells

            def serialize(self):
                return '{} {}'.format(self.value, self.units)

        class Model(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            attr1 = core.StringAttribute()
            attr2 = core.StringAttribute()
            quantity = core.OneToOneAttribute(Quantity, related_name='model')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'attr1', 'attr2')
                table_format = core.TableFormat.column

        filename = os.path.join(self.dirname, 'test.xlsx')
        writer_cls = get_writer('.xlsx')
        writer = writer_cls(filename)

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row([None, '!Id', 'm1']))
        ws.append(Row([None, '!Attr1', '1']))
        ws.append(Row([None, '!Attr2', '2']))
        ws.append(Row(['!Quantity', '!Value', 1.1]))
        ws.append(Row(['!Quantity', '!Units', 's']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        WorkbookReader().run(filename, models=[Model])

        wb = Workbook()
        wb['!Models'] = ws = Worksheet()
        wb['!Models'].append(Row(["!!ObjTables TableType='Data' ModelId='Model'"]))
        ws.append(Row([None, '!Id', 'm1']))
        ws.append(Row([None, '!Attr2', '2']))
        ws.append(Row([None, '!Attr1', '1']))
        ws.append(Row(['!Quantity', '!Value', 1.1]))
        ws.append(Row(['!Quantity', '!Units', 's']))
        writer.run(wb, style={
            Model.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(ValueError, (
            "The rows of worksheet '!Models' must be defined in this order:"
            "\n      A1: "
            "\n      B1: !Id"
            "\n      A2: "
            "\n      B2: !Attr1"
            "\n      A3: "
            "\n      B3: !Attr2"
            "\n      A4: !Quantity"
            "\n      B4: !Value"
            "\n      A5: !Quantity"
            "\n      B5: !Units"
        )):
            WorkbookReader().run(filename, models=[Model])
        WorkbookReader().run(filename, models=[Model], ignore_attribute_order=True)


class JsonTestCase(unittest.TestCase):
    def setUp(self):
        self.dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.dirname)

    def test_write_read(self):
        class AA(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            val = core.IntegerAttribute(min=0)

        class BB(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            aa = core.ManyToOneAttribute(AA, related_name='bbs')

        class CC(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            bbs = core.ManyToManyAttribute(BB, related_name='ccs')
            aas = core.ManyToManyAttribute(AA, related_name='ccs')

        aa_0 = AA(id='aa_0', val=1)
        aa_1 = AA(id='aa_1', val=2)
        aa_2 = AA(id='aa_2', val=3)

        bb_0_0 = aa_0.bbs.create(id='bb_0_0')
        bb_0_1 = aa_0.bbs.create(id='bb_0_1')

        cc_0_0_0 = bb_0_0.ccs.create(id='cc_0_0_0')
        cc_0_0_1 = bb_0_0.ccs.create(id='cc_0_0_1')
        cc_0_1_0 = bb_0_1.ccs.create(id='cc_0_1_0')
        cc_0_1_1 = bb_0_1.ccs.create(id='cc_0_1_1')

        cc_0_0_0.aas = [aa_0]
        cc_0_0_1.aas = [aa_1, aa_2]

        path = os.path.join(self.dirname, 'out.json')
        obj_tables.io.JsonWriter().run(path, aa_0)
        aa_0_2 = obj_tables.io.JsonReader().run(path, models=[AA])
        self.assertTrue(aa_0.is_equal(aa_0_2))

        obj_tables.io.JsonWriter().run(path, [aa_0, aa_1], models=AA)
        aas = obj_tables.io.JsonReader().run(path, models=[AA])
        aas.sort(key=lambda aa: aa.id)
        self.assertEqual(len(aas), 2)
        self.assertTrue(aa_0.is_equal(aas[0]))
        self.assertTrue(aa_1.is_equal(aas[1]))

        with open(path, 'r') as file:
            objs = json.load(file)
        objs[0]['val'] = -1
        with open(path, 'w') as file:
            json.dump(objs, file)
        with self.assertRaisesRegex(ValueError, 'fails to validate'):
            obj_tables.io.JsonReader().run(path, models=[AA])

        obj_tables.io.JsonWriter().run(path, aa_0, models=AA)
        aa_0_2 = obj_tables.io.JsonReader().run(path, models=AA)
        self.assertTrue(aa_0.is_equal(aa_0_2))

        obj_tables.io.JsonWriter().run(path, aa_0)
        aa_0_2 = obj_tables.io.JsonReader().run(path, models=AA)
        self.assertTrue(aa_0.is_equal(aa_0_2))
        aa_0_2 = obj_tables.io.JsonReader().run(path, models=AA, group_objects_by_model=True)
        self.assertEqual(list(aa_0_2.keys()), [AA])
        self.assertEqual(len(aa_0_2[AA]), 1)
        self.assertTrue(aa_0.is_equal(aa_0_2[AA][0]))

        obj_tables.io.JsonWriter().run(path, None)
        self.assertEqual(obj_tables.io.JsonReader().run(path), None)
        self.assertEqual(obj_tables.io.JsonReader().run(path, models=AA, group_objects_by_model=True), {})

        path = os.path.join(self.dirname, 'out.yml')
        obj_tables.io.JsonWriter().run(path, aa_0)
        aa_0_2 = obj_tables.io.JsonReader().run(path, models=[AA])
        self.assertTrue(aa_0.is_equal(aa_0_2))

        path = os.path.join(self.dirname, 'out.yml')
        obj_tables.io.Writer().run(path, aa_0)
        aa_0_2 = obj_tables.io.Reader().run(path, models=[AA])
        self.assertTrue(aa_0.is_equal(aa_0_2))

        path = os.path.join(self.dirname, 'out.yml')
        with self.assertWarnsRegex(obj_tables.io.IoWarning, 'has no effect'):
            obj_tables.io.JsonWriter().run(path, aa_0, include_all_attributes=False)

        path = os.path.join(self.dirname, 'out.abc')
        with self.assertRaisesRegex(ValueError, 'Unsupported format'):
            obj_tables.io.JsonWriter().run(path, aa_0)
        with self.assertRaisesRegex(ValueError, 'Unsupported format'):
            obj_tables.io.JsonReader().run(path, models=[AA])

        old_AA = AA

        class AA(core.Model):
            id = core.StringAttribute()
        path = os.path.join(self.dirname, 'out.yml')
        with self.assertRaisesRegex(ValueError, 'Model names must be unique to decode objects'):
            obj_tables.io.JsonWriter().run(path, aa_0, models=[AA, old_AA])
        with self.assertRaisesRegex(ValueError, 'Model names must be unique to decode objects'):
            obj_tables.io.JsonReader().run(path, models=[AA, old_AA])

    def test_convert(self):
        root = MainRoot(id='root', name=u'\u20ac')
        nodes = [
            Node(root=root, id='node_0', val1=1, val2=2),
            Node(root=root, id='node_1', val1=3, val2=4),
            Node(root=root, id='node_2', val1=5, val2=6),
        ]
        leaves = [
            Leaf(nodes=[nodes[0]], id='leaf_0_0', val1=7, val2=8),
            Leaf(nodes=[nodes[0]], id='leaf_0_1', val1=9, val2=10),
            Leaf(nodes=[nodes[1]], id='leaf_1_0', val1=11, val2=12),
            Leaf(nodes=[nodes[1]], id='leaf_1_1', val1=13, val2=14),
            Leaf(nodes=[nodes[2]], id='leaf_2_0', val1=15, val2=16),
            Leaf(nodes=[nodes[2]], id='leaf_2_1', val1=17, val2=18),
        ]
        leaves[0].onetomany_rows = [OneToManyRow(id='row_0_0'), OneToManyRow(id='row_0_1')]
        leaves[1].onetomany_rows = [OneToManyRow(id='row_1_0'), OneToManyRow(id='row_1_1')]
        leaves[2].onetomany_rows = [OneToManyRow(id='row_2_0'), OneToManyRow(id='row_2_1')]
        leaves[3].onetomany_rows = [OneToManyRow(id='row_3_0'), OneToManyRow(id='row_3_1')]
        leaves[4].onetomany_rows = [OneToManyRow(id='row_4_0'), OneToManyRow(id='row_4_1')]
        leaves[5].onetomany_rows = [OneToManyRow(id='row_5_0'), OneToManyRow(id='row_5_1')]

        leaves[0].onetomany_inlines = [OneToManyInline(id='inline_0_0'), OneToManyInline(id='inline_0_1')]
        leaves[1].onetomany_inlines = [OneToManyInline(id='inline_1_0'), OneToManyInline(id='inline_1_1')]
        leaves[2].onetomany_inlines = [OneToManyInline(id='inline_2_0'), OneToManyInline(id='inline_2_1')]
        leaves[3].onetomany_inlines = [OneToManyInline(id='inline_3_0'), OneToManyInline(id='inline_3_1')]
        leaves[4].onetomany_inlines = [OneToManyInline(id='inline_4_0'), OneToManyInline(id='inline_4_1')]
        leaves[5].onetomany_inlines = [OneToManyInline(id='inline_5_0'), OneToManyInline(id='inline_5_1')]

        filename_1_xlsx = os.path.join(self.dirname, 'test_1.xlsx')
        filename_2_json = os.path.join(self.dirname, 'test_2.json')
        filename_3_xlsx = os.path.join(self.dirname, 'test_3.xlsx')

        models = [MainRoot, Node, Leaf, OneToManyRow]

        obj_tables.io.Writer.get_writer(filename_1_xlsx)().run(filename_1_xlsx, [root], models=models)

        # convert xlsx --> json
        convert(filename_1_xlsx, filename_2_json, models=models)

        objects2 = obj_tables.io.Reader.get_reader(filename_2_json)().run(filename_2_json, models=models,
                                                                          group_objects_by_model=True)
        self.assertEqual(len(objects2[MainRoot]), 1)
        root2 = objects2[MainRoot][0]
        self.assertTrue(root.is_equal(root2))

        # convert json --> xlsx
        convert(filename_2_json, filename_3_xlsx, models=models)

        objects2 = obj_tables.io.Reader.get_reader(filename_3_xlsx)().run(filename_3_xlsx, models=models,
                                                                          group_objects_by_model=True)
        self.assertEqual(len(objects2[MainRoot]), 1)
        root2 = objects2[MainRoot][0]
        self.assertTrue(root.is_equal(root2))

    def test_write_invalid(self):
        class Node(core.Model):
            id = core.StringAttribute(min_length=3)

        node = Node(id='a')
        self.assertNotEqual(core.Validator().run(node), None)
        with pytest.warns(IoWarning, match='objects are not valid'):
            writer = obj_tables.io.JsonWriter()
            filename = os.path.join(self.dirname, 'test.json')
            writer.run(filename, [node], models=[Node])


class InlineJsonTestCase(unittest.TestCase):
    def setUp(self):
        self.dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.dirname)

    def test_no_primary(self):
        class OtherGrandChild(core.Model):
            name = core.StringAttribute()

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell

        class GrandChild(core.Model):
            name = core.StringAttribute()
            sibling = core.OneToOneAttribute(OtherGrandChild, related_name='sibling')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell

        class Child(core.Model):
            name = core.StringAttribute()
            children = core.OneToManyAttribute(GrandChild, related_name='parent')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell

        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            children = core.OneToManyAttribute(Child, related_name='parent')

        p = Parent(id='p')
        c0 = p.children.create(name='c0')
        c1 = p.children.create(name='c1')
        g00 = c0.children.create(name='g00')
        g01 = c0.children.create(name='g01')
        g10 = c1.children.create(name='g10')
        g11 = c1.children.create(name='g11')
        g00.sibling = OtherGrandChild(name='o00')
        g11.sibling = OtherGrandChild(name='o11')

        path = os.path.join(self.dirname, 'test.xlsx')
        obj_tables.io.WorkbookWriter().run(path, p)
        objs2 = obj_tables.io.WorkbookReader().run(path, models=Parent)
        self.assertEqual(list(objs2.keys()), [Parent])
        self.assertEqual(len(objs2[Parent]), 1)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

    def test_primary_attribute(self):
        class GrandChild(core.Model):
            name = core.StringAttribute(primary=True, unique=True)

        class Child(core.Model):
            name = core.StringAttribute()
            child = core.ManyToOneAttribute(GrandChild, related_name='parents')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell

        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            children = core.OneToManyAttribute(Child, related_name='parent')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'children')

        p = Parent(id='p')
        c0 = p.children.create(name='c0')
        c1 = p.children.create(name='c1')
        g00 = c0.child = GrandChild(name='g00')
        g01 = c0.child = GrandChild(name='g01')
        g10 = c1.child = GrandChild(name='g10')
        g11 = c1.child = GrandChild(name='g11')

        path = os.path.join(self.dirname, 'test.xlsx')
        obj_tables.io.WorkbookWriter().run(path, p)
        objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)
        self.assertEqual(set(objs2.keys()), set([Parent, GrandChild]))
        self.assertEqual(len(objs2[Parent]), 1)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

        # test exception
        wb = read_workbook(path)
        wb['!Parents'][2][1] = ']'
        write_workbook(path, wb, style={
            Parent.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            Child.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            GrandChild.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(Exception, 'test.xlsx:!Parents:B2'):
            objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)

    def test_one_to_one(self):
        class GrandChild(core.Model):
            name = core.StringAttribute(primary=True, unique=True)

        class Child(core.Model):
            name = core.StringAttribute()
            children = core.OneToManyAttribute(GrandChild, related_name='parent')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell

        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            child = core.OneToOneAttribute(Child, related_name='parent')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'child')

        p = Parent(id='p')
        c0 = p.child = Child(name='c0')
        c1 = p.child = Child(name='c1')
        g00 = c0.children.create(name='g00')
        g01 = c0.children.create(name='g01')
        g10 = c1.children.create(name='g10')
        g11 = c1.children.create(name='g11')

        path = os.path.join(self.dirname, 'test.xlsx')
        obj_tables.io.WorkbookWriter().run(path, p)
        objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)
        self.assertEqual(set(objs2.keys()), set([Parent, GrandChild]))
        self.assertEqual(len(objs2[Parent]), 1)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

        # test exception
        wb = read_workbook(path)
        wb['!Parents'][2][1] = ']'
        write_workbook(path, wb, style={
            Parent.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            Child.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            GrandChild.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(Exception, 'test.xlsx:!Parents:B2'):
            objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)

    def test_many_to_one(self):
        class GrandChild(core.Model):
            name = core.StringAttribute(primary=True, unique=True)

        class Child(core.Model):
            name = core.StringAttribute()
            children = core.ManyToManyAttribute(GrandChild, related_name='parents')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell

        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            child = core.ManyToOneAttribute(Child, related_name='parents')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'child')

        p = Parent(id='p')
        c0 = p.child = Child(name='c0')
        c1 = p.child = Child(name='c1')
        g00 = c0.children.create(name='g00')
        g01 = c0.children.create(name='g01')
        g10 = c1.children.create(name='g10')
        g11 = c1.children.create(name='g11')

        path = os.path.join(self.dirname, 'test.xlsx')
        obj_tables.io.WorkbookWriter().run(path, p)
        objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)
        self.assertEqual(set(objs2.keys()), set([Parent, GrandChild]))
        self.assertEqual(len(objs2[Parent]), 1)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

        # test exception
        wb = read_workbook(path)
        wb['!Parents'][2][1] = ']'
        write_workbook(path, wb, style={
            Parent.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            Child.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            GrandChild.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(Exception, 'test.xlsx:!Parents:B2'):
            objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)

    def test_many_to_many(self):
        class GrandChild(core.Model):
            name = core.StringAttribute(primary=True, unique=True)

        class Child(core.Model):
            name = core.StringAttribute()
            child = core.OneToOneAttribute(GrandChild, related_name='parent')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell

        class Parent(core.Model):
            id = core.StringAttribute(primary=True, unique=True)
            children = core.ManyToManyAttribute(Child, related_name='parents')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'children')

        p = Parent(id='p')
        c0 = p.children.create(name='c0')
        c1 = p.children.create(name='c1')
        g00 = c0.child = GrandChild(name='g00')
        g01 = c0.child = GrandChild(name='g01')
        g10 = c1.child = GrandChild(name='g10')
        g11 = c1.child = GrandChild(name='g11')

        path = os.path.join(self.dirname, 'test.xlsx')
        obj_tables.io.WorkbookWriter().run(path, p)
        objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)
        self.assertEqual(set(objs2.keys()), set([Parent, GrandChild]))
        self.assertEqual(len(objs2[Parent]), 1)
        p2 = objs2[Parent][0]
        self.assertTrue(p.is_equal(p2))

        # test exception
        wb = read_workbook(path)
        wb['!Parents'][2][1] = ']'
        write_workbook(path, wb, style={
            Parent.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            Child.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
            GrandChild.Meta.verbose_name_plural: WorksheetStyle(extra_rows=0, extra_columns=0),
        })
        with self.assertRaisesRegex(Exception, 'test.xlsx:!Parents:B2'):
            objs2 = obj_tables.io.WorkbookReader().run(path, models=[Parent, GrandChild], ignore_sheet_order=True)


class UtilsTestCase(unittest.TestCase):
    def test_get_writer(self):
        self.assertEqual(obj_tables.io.Writer.get_writer('test.csv'), obj_tables.io.WorkbookWriter)
        self.assertEqual(obj_tables.io.Writer.get_writer('test.tsv'), obj_tables.io.WorkbookWriter)
        self.assertEqual(obj_tables.io.Writer.get_writer('test.xlsx'), obj_tables.io.WorkbookWriter)
        self.assertEqual(obj_tables.io.Writer.get_writer('test.json'), obj_tables.io.JsonWriter)
        self.assertEqual(obj_tables.io.Writer.get_writer('test.yaml'), obj_tables.io.JsonWriter)
        self.assertEqual(obj_tables.io.Writer.get_writer('test.yml'), obj_tables.io.JsonWriter)

        with self.assertRaises(ValueError):
            obj_tables.io.Writer.get_writer('test.abc')

    def test_get_reader(self):
        self.assertEqual(obj_tables.io.Reader.get_reader('test.csv'), obj_tables.io.WorkbookReader)
        self.assertEqual(obj_tables.io.Reader.get_reader('test.tsv'), obj_tables.io.WorkbookReader)
        self.assertEqual(obj_tables.io.Reader.get_reader('test.xlsx'), obj_tables.io.WorkbookReader)
        self.assertEqual(obj_tables.io.Reader.get_reader('test.json'), obj_tables.io.JsonReader)
        self.assertEqual(obj_tables.io.Reader.get_reader('test.yaml'), obj_tables.io.JsonReader)
        self.assertEqual(obj_tables.io.Reader.get_reader('test.yml'), obj_tables.io.JsonReader)

        with self.assertRaises(ValueError):
            obj_tables.io.Reader.get_reader('test.abc')

    def test_get_ordered_attributes(self):
        class Root(core.Model):
            label = core.StringAttribute(primary=True, unique=True)

        class Leaf(core.Model):
            root = core.ManyToOneAttribute(Root, related_name='leaves')
            id = core.StringAttribute(primary=True)
            name = core.StringAttribute()

            class Meta(core.Model.Meta):
                attribute_order = ('id', )

        class UnrootedLeaf(Leaf):
            name = core.StringAttribute()
            root2 = core.ManyToOneAttribute(Root, related_name='leaves2')
            id2 = core.StringAttribute()
            name2 = core.StringAttribute()
            float2 = core.FloatAttribute()
            float3 = core.FloatAttribute()
            enum2 = core.StringAttribute()
            enum3 = core.StringAttribute()
            multi_word_name = core.StringAttribute()

        class Leaf3(UnrootedLeaf):
            class Meta(core.Model.Meta):
                attribute_order = ('id2', 'name2', )

        # all attributes
        root_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(Root))
        leaf_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(Leaf))
        unrooted_leaf_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(UnrootedLeaf))
        leaf3_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(Leaf3))

        self.assertEqual(set(root_attrs), set(Root.Meta.attributes.keys()))
        self.assertEqual(set(leaf_attrs), set(Leaf.Meta.attributes.keys()))
        self.assertEqual(set(unrooted_leaf_attrs), set(UnrootedLeaf.Meta.attributes.keys()))
        self.assertEqual(set(leaf3_attrs), set(Leaf3.Meta.attributes.keys()))

        self.assertEqual(root_attrs, ('label', ))
        self.assertEqual(leaf_attrs, ('id', 'name', 'root'))
        self.assertEqual(unrooted_leaf_attrs, (
            'id',
            'enum2', 'enum3', 'float2', 'float3', 'id2', 'multi_word_name', 'name', 'name2', 'root', 'root2', ))
        self.assertEqual(leaf3_attrs, (
            'id2', 'name2',
            'enum2', 'enum3', 'float2', 'float3', 'id', 'multi_word_name', 'name', 'root', 'root2', ))

        # only explicitly defined attributes
        root_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(Root, include_all_attributes=False))
        leaf_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(Leaf, include_all_attributes=False))
        unrooted_leaf_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(
            UnrootedLeaf, include_all_attributes=False))
        leaf3_attrs = tuple(attr.name for attr in obj_tables.io.get_ordered_attributes(Leaf3, include_all_attributes=False))

        self.assertLessEqual(set(root_attrs), set(Root.Meta.attributes.keys()))
        self.assertLessEqual(set(leaf_attrs), set(Leaf.Meta.attributes.keys()))
        self.assertLessEqual(set(unrooted_leaf_attrs), set(UnrootedLeaf.Meta.attributes.keys()))
        self.assertLessEqual(set(leaf3_attrs), set(Leaf3.Meta.attributes.keys()))

        self.assertEqual(root_attrs, ())
        self.assertEqual(leaf_attrs, ('id',))
        self.assertEqual(unrooted_leaf_attrs, ('id',))
        self.assertEqual(leaf3_attrs, ('id2', 'name2',))

    def test_get_ordered_attributes_error(self):
        class Unit(core.Model):
            id = core.SlugAttribute()

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.cell
                attribute_order = ('id',)

            def serialize(self):
                return self.id

        class Quantity(core.Model):
            value = core.FloatAttribute()
            unit = core.OneToOneAttribute(Unit, related_name='quantity')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.multiple_cells
                attribute_order = ('value', 'unit')

            def serialize(self):
                return '{} {}'.format(self.value, self.unit.id)

        class Model(core.Model):
            quantity = core.OneToOneAttribute(Quantity, related_name='model')

        obj_tables.io.get_ordered_attributes(Model)
        obj_tables.io.get_ordered_attributes(Quantity)
        obj_tables.io.get_ordered_attributes(Unit)

        class Unit(core.Model):
            id = core.SlugAttribute()

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.multiple_cells
                attribute_order = ('id',)

            def serialize(self):
                return self.id

            @classmethod
            def deserialize(cls, value, objects):
                return cls(id=value)

        class Quantity(core.Model):
            value = core.FloatAttribute()
            unit = core.OneToOneAttribute(Unit, related_name='quantity')

            class Meta(core.Model.Meta):
                table_format = core.TableFormat.multiple_cells
                attribute_order = ('value', 'unit')

            def serialize(self):
                return '{} {}'.format(self.value, self.unit.id)

        class Model(core.Model):
            quantity = core.OneToOneAttribute(Quantity, related_name='model')

        obj_tables.io.get_ordered_attributes(Model)
        with self.assertRaisesRegex(ValueError, 'cannot have relationships'):
            obj_tables.io.get_ordered_attributes(Quantity)
        obj_tables.io.get_ordered_attributes(Unit)


class ExcelValidationTestCase(unittest.TestCase):
    def setUp(self):
        self.dirname = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.dirname)

    def test(self):
        class TestEnum(enum.Enum):
            val1 = 0
            val2 = 1

        class TestChild1(core.Model):
            id = core.SlugAttribute(unique=True, primary=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

        class TestChild2(core.Model):
            id = core.SlugAttribute(unique=True, primary=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)
                table_format = core.TableFormat.column

        sbo_ontotology = pronto.Ontology('tests/fixtures/SBO.obo')
        unit_registry = pint.UnitRegistry()

        class Parameter1(core.Model):
            id = core.SlugAttribute()
            value = core.FloatAttribute()
            units = units.UnitAttribute(unit_registry)

            class Meta(core.Model.Meta, obj_tables.expression.ExpressionStaticTermMeta):
                expression_term_value = 'value'
                expression_term_units = 'units'

        class Parameter2(core.Model):
            id = core.SlugAttribute()
            value = core.FloatAttribute()
            units = units.UnitAttribute(unit_registry)

            class Meta(core.Model.Meta, obj_tables.expression.ExpressionStaticTermMeta):
                expression_term_value = 'value'
                expression_term_units = 'units'

        class TestParentExpression1(core.Model, obj_tables.expression.Expression):
            expression = core.StringAttribute()

            class Meta(core.Model.Meta, obj_tables.expression.Expression.Meta):
                expression_term_models = ()
                expression_unit_registry = unit_registry

            def serialize(self): return obj_tables.expression.Expression.serialize(self)

            @classmethod
            def deserialize(cls, value, objects): return obj_tables.expression.Expression.deserialize(cls, value, objects)

            def validate(self): return obj_tables.expression.Expression.validate(self, self.parent_sub_function)

        class TestParentExpression2(core.Model, obj_tables.expression.Expression):
            expression = core.StringAttribute()
            parameters_1 = core.ManyToManyAttribute(Parameter1, related_name='test_parent_expressions_2')

            class Meta(core.Model.Meta, obj_tables.expression.Expression.Meta):
                expression_term_models = ('Parameter1',)
                expression_unit_registry = unit_registry

            def serialize(self): return obj_tables.expression.Expression.serialize(self)

            @classmethod
            def deserialize(cls, value, objects): return obj_tables.expression.Expression.deserialize(cls, value, objects)

            def validate(self): return obj_tables.expression.Expression.validate(self, self.parent_sub_function)

        class TestParentExpression3(core.Model, obj_tables.expression.Expression):
            expression = core.StringAttribute()
            parameters_1 = core.ManyToManyAttribute(Parameter1, related_name='test_parent_expressions_3')
            parameters_2 = core.ManyToManyAttribute(Parameter2, related_name='test_parent_expressions_3')

            class Meta(core.Model.Meta, obj_tables.expression.Expression.Meta):
                expression_term_models = ('Parameter1', 'Parameter2')
                expression_unit_registry = unit_registry
                expression_is_linear = True

            def serialize(self): return obj_tables.expression.Expression.serialize(self)

            @classmethod
            def deserialize(cls, value, objects): return obj_tables.expression.Expression.deserialize(cls, value, objects)

            def validate(self): return obj_tables.expression.Expression.validate(self, self.parent_sub_function)

        class TestParent(core.Model):
            id = core.SlugAttribute(unique=True, primary=True)
            enum_attr_1 = core.EnumAttribute(TestEnum, default_cleaned_value=TestEnum.val1)
            enum_attr_2 = core.EnumAttribute(TestEnum, none=True, unique=True)
            bool_attr = core.BooleanAttribute(default_cleaned_value=True)
            float_attr_1 = core.FloatAttribute(default_cleaned_value=4.)
            float_attr_2 = core.FloatAttribute(default_cleaned_value=4., min=-1.)
            float_attr_3 = core.FloatAttribute(default_cleaned_value=4., max=-1.)
            float_attr_4 = core.FloatAttribute(default_cleaned_value=4., nan=False, min=-1., max=1., unique=True)
            pos_float_attr_1 = core.PositiveFloatAttribute(default_cleaned_value=5.)
            pos_float_attr_2 = core.PositiveFloatAttribute(default_cleaned_value=5., nan=False, max=10., unique=True)
            int_attr_1 = core.IntegerAttribute(default_cleaned_value=2)
            int_attr_2 = core.IntegerAttribute(default_cleaned_value=2, min=-1)
            int_attr_3 = core.IntegerAttribute(default_cleaned_value=2, max=1)
            int_attr_4 = core.IntegerAttribute(default_cleaned_value=2, min=-2, max=2, unique=True)
            pos_int_attr_1 = core.PositiveIntegerAttribute(default_cleaned_value=3)
            pos_int_attr_2 = core.PositiveIntegerAttribute(default_cleaned_value=3, max=10, unique=True)

            str_attr_1 = core.StringAttribute(default_cleaned_value='default val', description='Enter a string')
            str_attr_2 = core.StringAttribute(min_length=1, max_length=None)
            str_attr_3 = core.StringAttribute(min_length=0, max_length=10)
            str_attr_4 = core.StringAttribute(min_length=1, max_length=10, unique=True)

            date_attr = core.DateAttribute(default_cleaned_value=datetime.date(2000, 1, 2), unique=True)
            time_attr = core.TimeAttribute(default_cleaned_value=datetime.time(10, 1, 2), unique=True)
            date_time_attr = core.DateTimeAttribute(default_cleaned_value=datetime.datetime(2001, 2, 3, 11, 3, 4), unique=True)
            one_to_one_attr_1 = core.OneToOneAttribute(TestChild1, related_name='parent_1')
            many_to_one_attr_1 = core.ManyToOneAttribute(TestChild1, related_name='parents_2')
            one_to_many_attr_1 = core.OneToManyAttribute(TestChild1, related_name='parent_3')
            many_to_many_attr_1 = core.ManyToManyAttribute(TestChild1, related_name='parents_4')
            one_to_one_attr_2 = core.OneToOneAttribute(TestChild2, related_name='parent_1',
                                                       min_related=1, default_cleaned_value=lambda: TestChild2(id='child_b_1'))
            many_to_one_attr_2 = core.ManyToOneAttribute(TestChild2, related_name='parents_2',
                                                         min_related=1, default_cleaned_value=lambda: TestChild2(id='child_b_2'))
            one_to_many_attr_2 = core.OneToManyAttribute(TestChild2, related_name='parent_3',
                                                         min_related=1, default_cleaned_value=lambda: [TestChild2(id='child_b_3')])
            many_to_many_attr_2 = core.ManyToManyAttribute(TestChild2, related_name='parents_4',
                                                           min_related=1, default_cleaned_value=lambda: [TestChild2(id='child_b_4')])
            formula_attr = chem.EmpiricalFormulaAttribute(unique=True)
            onto_attr_1 = ontology.OntologyAttribute(sbo_ontotology,
                                                     namespace='SBO',
                                                     terms=sbo_ontotology['SBO:0000474'].rchildren(),
                                                     default_cleaned_value=sbo_ontotology['SBO:0000475'])
            onto_attr_2 = ontology.OntologyAttribute(sbo_ontotology,
                                                     namespace='SBO',
                                                     terms=sbo_ontotology['SBO:0000474'].rchildren(),
                                                     default_cleaned_value=sbo_ontotology['SBO:0000475'],
                                                     unique=True, none=False)
            onto_attr_3 = ontology.OntologyAttribute(sbo_ontotology,
                                                     namespace='SBO',
                                                     default_cleaned_value=sbo_ontotology['SBO:0000475'])
            onto_attr_4 = ontology.OntologyAttribute(sbo_ontotology,
                                                     namespace='SBO',
                                                     default_cleaned_value=sbo_ontotology['SBO:0000475'],
                                                     unique=True, none=False)
            units_attr_1 = units.UnitAttribute(unit_registry, choices=(
                unit_registry.parse_units('g'),
                unit_registry.parse_units('l'),
            ), default_cleaned_value=unit_registry.parse_units('g'))
            units_attr_2 = units.UnitAttribute(unit_registry, choices=(
                unit_registry.parse_units('g'),
                unit_registry.parse_units('l'),
            ), default_cleaned_value=unit_registry.parse_units('g'), unique=True, none=False)
            units_attr_3 = units.UnitAttribute(unit_registry, default_cleaned_value=unit_registry.parse_units('g'))
            units_attr_4 = units.UnitAttribute(
                unit_registry, default_cleaned_value=unit_registry.parse_units('g'), unique=True, none=False)
            expression_attr_1 = obj_tables.expression.ExpressionOneToOneAttribute(TestParentExpression1, related_name='test_parent')
            expression_attr_2 = obj_tables.expression.ExpressionOneToOneAttribute(TestParentExpression2, related_name='test_parent')
            expression_attr_3 = obj_tables.expression.ExpressionOneToOneAttribute(TestParentExpression3, related_name='test_parent')
            expression_attr_4 = obj_tables.expression.ExpressionManyToOneAttribute(TestParentExpression1, related_name='test_parents')
            expression_attr_5 = obj_tables.expression.ExpressionManyToOneAttribute(TestParentExpression2, related_name='test_parents')
            expression_attr_6 = obj_tables.expression.ExpressionManyToOneAttribute(TestParentExpression3, related_name='test_parents')

            class Meta(core.Model.Meta):
                attribute_order = ('id', 'enum_attr_1', 'enum_attr_2',
                                   'bool_attr',
                                   'float_attr_1', 'float_attr_2', 'float_attr_3', 'float_attr_4',
                                   'pos_float_attr_1', 'pos_float_attr_2',
                                   'int_attr_1', 'int_attr_2', 'int_attr_3', 'int_attr_4',
                                   'pos_int_attr_1', 'pos_int_attr_2',
                                   'str_attr_1', 'str_attr_2', 'str_attr_3', 'str_attr_4',
                                   'date_attr', 'time_attr', 'date_time_attr',
                                   'one_to_one_attr_1', 'many_to_one_attr_1', 'one_to_many_attr_1', 'many_to_many_attr_1',
                                   'one_to_one_attr_2', 'many_to_one_attr_2', 'one_to_many_attr_2', 'many_to_many_attr_2',
                                   'formula_attr',
                                   'onto_attr_1', 'onto_attr_2', 'onto_attr_3', 'onto_attr_4',
                                   'units_attr_1', 'units_attr_2', 'units_attr_3', 'units_attr_4',
                                   'expression_attr_1', 'expression_attr_2', 'expression_attr_3',
                                   'expression_attr_4', 'expression_attr_5', 'expression_attr_6',
                                   )

        for attr in TestParent.Meta.attributes.values():
            attr.description = 'A helpful description'

        objects = [
            TestParent(id='parent_a',
                       enum_attr_2=TestEnum.val1,
                       float_attr_4=-0.5,
                       pos_float_attr_2=5.,
                       int_attr_1=1, int_attr_2=1, int_attr_3=-1, int_attr_4=1,
                       pos_int_attr_1=1, pos_int_attr_2=3,
                       str_attr_2='a2', str_attr_4='a4',
                       date_attr=datetime.date(2001, 1, 1),
                       time_attr=datetime.time(11, 0, 0),
                       date_time_attr=datetime.datetime(2001, 1, 1, 11, 0, 0),
                       formula_attr=wc_utils.util.chem.EmpiricalFormula('H2O'),
                       onto_attr_2=sbo_ontotology['SBO:0000475'],
                       onto_attr_4=sbo_ontotology['SBO:0000475'],
                       units_attr_2=unit_registry.parse_units('g'),
                       units_attr_4=unit_registry.parse_units('g')),
            TestParent(id='parent_b',
                       enum_attr_2=TestEnum.val2,
                       float_attr_4=0.5,
                       pos_float_attr_2=8.,
                       int_attr_1=1, int_attr_2=1, int_attr_3=-1, int_attr_4=2,
                       pos_int_attr_1=1, pos_int_attr_2=5,
                       str_attr_2='b2', str_attr_4='b4',
                       date_attr=datetime.date(2001, 1, 2),
                       time_attr=datetime.time(12, 0, 0),
                       date_time_attr=datetime.datetime(2001, 1, 2, 12, 0, 0),
                       formula_attr=wc_utils.util.chem.EmpiricalFormula('CO2'),
                       onto_attr_2=sbo_ontotology['SBO:0000487'],
                       onto_attr_4=sbo_ontotology['SBO:0000487'],
                       units_attr_2=unit_registry.parse_units('l'),
                       units_attr_4=unit_registry.parse_units('l')),
            TestChild1(id='child_1_a'),
            TestChild1(id='child_1_b'),
            TestChild2(id='child_2_a'),
            TestChild2(id='child_2_b'),
        ]

        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, objects, models=[TestParent, TestChild1, TestChild2])

    def test_no_primary_attr(self):
        class TestChild1(core.Model):
            id = core.StringAttribute(unique=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)

            def serialize(self):
                return self.id

        class TestChild2(core.Model):
            id = core.StringAttribute(unique=True)

            class Meta(core.Model.Meta):
                attribute_order = ('id',)
                table_format = core.TableFormat.column

            def serialize(self):
                return self.id

        class OneToOneAttribute(core.OneToOneAttribute):
            def serialize(self, value, encoded=None):
                if value:
                    return value.serialize()
                else:
                    return None

            def deserialize(self, value, objects, decoded=None): pass

        class OneToManyAttribute(core.OneToManyAttribute):
            def serialize(self, values, encoded=None):
                return ', '.join([value.serialize() for value in values])

            def deserialize(self, value, objects, decoded=None): pass

        class ManyToOneAttribute(core.ManyToOneAttribute):
            def serialize(self, value, encoded=None):
                if value:
                    return value.serialize()
                else:
                    return None

            def deserialize(self, value, objects, decoded=None): pass

        class ManyToManyAttribute(core.ManyToManyAttribute):
            def serialize(self, values, encoded=None):
                return ', '.join([value.serialize() for value in values])

            def deserialize(self, value, objects, decoded=None): pass

        class TestParent(core.Model):
            id = core.StringAttribute(unique=True)
            one_to_one_attr_1 = OneToOneAttribute(TestChild1, related_name='parent_1')
            many_to_one_attr_1 = ManyToOneAttribute(TestChild1, related_name='parents_2')
            one_to_many_attr_1 = OneToManyAttribute(TestChild1, related_name='parent_3')
            many_to_many_attr_1 = ManyToManyAttribute(TestChild1, related_name='parents_4')
            one_to_one_attr_2 = OneToOneAttribute(TestChild2, related_name='parent_1',
                                                  min_related=1, default_cleaned_value=lambda: TestChild2(id='child_b_1'))
            many_to_one_attr_2 = ManyToOneAttribute(TestChild2, related_name='parents_2',
                                                    min_related=1, default_cleaned_value=lambda: TestChild2(id='child_b_2'))
            one_to_many_attr_2 = OneToManyAttribute(TestChild2, related_name='parent_3',
                                                    min_related=1, default_cleaned_value=lambda: [TestChild2(id='child_b_3')])
            many_to_many_attr_2 = ManyToManyAttribute(TestChild2, related_name='parents_4',
                                                      min_related=1, default_cleaned_value=lambda: [TestChild2(id='child_b_4')])

            class Meta(core.Model.Meta):
                attribute_order = ('id',
                                   'one_to_one_attr_1', 'many_to_one_attr_1', 'one_to_many_attr_1', 'many_to_many_attr_1',
                                   'one_to_one_attr_2', 'many_to_one_attr_2', 'one_to_many_attr_2', 'many_to_many_attr_2',
                                   )

            def serialize(self):
                return self.id

        for attr in TestParent.Meta.attributes.values():
            attr.description = 'A helpful description'

        objects = [
            TestParent(id='parent_a'),
            TestParent(id='parent_b'),
            TestChild1(id='child_1_a'),
            TestChild1(id='child_1_b'),
            TestChild2(id='child_2_a'),
            TestChild2(id='child_2_b'),
        ]

        filename = os.path.join(self.dirname, 'test.xlsx')
        WorkbookWriter().run(filename, objects, models=[TestParent, TestChild1, TestChild2])
