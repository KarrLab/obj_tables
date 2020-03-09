""" Migration to ObjTables format as of 2020-03-09

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2020-03-09
:Copyright: 2020, Karr Lab
:License: MIT
"""

import obj_tables
import openpyxl
import re
import stringcase


def transform(filename):
    # read
    wb = openpyxl.load_workbook(filename=filename)

    for ws in wb:
        if not ws.title.startswith('!'):
            continue

        if isinstance(ws.cell(1, 1).value, str) and ws.cell(1, 1).value.startswith('!!'):
            matches = re.findall(r" +(.*?)=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                ws.cell(1, 1).value)
            heading, _, _ = ws.cell(1, 1).value.partition(' ')
            for key, val, _, _ in matches:
                heading += ' {}={}'.format(stringcase.camelcase(key), val)
            ws.cell(1, 1).value = heading

        if isinstance(ws.cell(2, 1).value, str) and ws.cell(2, 1).value.startswith('!!'):
            matches = re.findall(r" +(.*?)=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                ws.cell(2, 1).value)
            heading, _, _ = ws.cell(2, 1).value.partition(' ')
            for key, val, _, _ in matches:
                heading += ' {}={}'.format(stringcase.camelcase(key), val)
            ws.cell(2, 1).value = heading

    # save
    wb.save(filename)
